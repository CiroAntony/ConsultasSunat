"""
=====================================================================
  CONSULTA MASIVA SUNAT  v2.1
  ============================
  Herramienta de escritorio para consultar masivamente el portal
  e-consultaruc.sunat.gob.pe usando automatización de navegador.

  FLUJO GENERAL DE EJECUCIÓN:
  ─────────────────────────────────────────────────────────────────
  1.  El usuario abre la app (App.__init__ → _build_ui → _check_deps)
  2.  Selecciona un archivo Excel con la hoja "ruc_unificado" (col A)
  3.  Pulsa INICIAR → App._iniciar() lanza App._ejecutar() en hilo
  4.  _ejecutar() carga RUCs, lee cache SQLite y llena una Queue
  5.  Se lanzan NUM_WORKERS hilos, cada uno corre worker_procesar_rucs()
  6.  Cada worker instancia Playwright/Chromium y consulta 3 etapas:
        a. consultar_ruc_principal()   → datos generales del contribuyente
        b. consultar_deuda_coactiva()  → deuda en cobranza coactiva
        c. consultar_trabajadores()    → histórico de declaraciones
  7.  Los resultados se persisten en SQLite y se muestran en la tabla
  8.  Al finalizar (o cancelar) se puede exportar a Excel formateado

  MÓDULOS / SECCIONES DEL ARCHIVO (en orden de definición):
  ─────────────────────────────────────────────────────────────────
  1  Imports y detección de dependencias opcionales
  2  Constantes globales (URLs, tiempos, límites, user-agents)
  3  Rate limiter global (_RateLimiter)
  4  Utilidades puras (helpers sin estado)
  5  Persistencia SQLite (cache de resultados)
  6  Playwright — gestión del driver (abrir/cerrar Chromium)
  7  Excepciones personalizadas
  8  Capa de scraping — 3 funciones de consulta SUNAT
  9  Orquestación de consulta completa (3 etapas + armado de fila)
  10 Worker (hilo productor de resultados)
  11 Exportación a Excel (openpyxl)
  12 Interfaz gráfica Tkinter (App)
  13 Punto de entrada __main__

  Dependencias:
    pip install playwright openpyxl
    playwright install chromium
=====================================================================
"""

# ══════════════════════════════════════════════════════════════════
#  1  IMPORTS Y DETECCIÓN DE DEPENDENCIAS
#  ─────────────────────────────────────────────────────────────────
#  Se importan primero los módulos de la stdlib (siempre disponibles)
#  y luego los de terceros (playwright, openpyxl).  Si alguno falta,
#  DEPS_OK = False y la UI informará al usuario sin crashear.
# ══════════════════════════════════════════════════════════════════

# ── Stdlib ────────────────────────────────────────────────────────
import time          # sleep, time.time para temporizadores
import re            # limpieza de texto con expresiones regulares
import os            # rutas de archivo, directorios
import threading     # hilos para los workers y locks
import queue         # colas de comunicación inter-hilo (thread-safe)
import random        # pausas y user-agents aleatorios (evasión WAF)
import sqlite3       # persistencia local de resultados (cache)
import json          # serialización de dicts a texto para SQLite
import tkinter as tk                       # GUI principal
from tkinter import filedialog, messagebox, ttk  # diálogos y widgets
from datetime import datetime              # timestamp en export/cache
from collections import deque             # ventana deslizante para ETA

# ── Terceros opcionales ───────────────────────────────────────────
# Cada bloque try/except acumula los módulos faltantes en MISSING.
DEPS_OK = True   # Se pone en False si falta cualquier dependencia
MISSING = []     # Lista de mensajes de error para mostrar en UI

try:
    from playwright.sync_api import (
        sync_playwright,
        TimeoutError as PlaywrightTimeoutError,
        Error as PlaywrightError,
    )
except ImportError:
    DEPS_OK = False
    MISSING.append("playwright  (pip install playwright && playwright install chromium)")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    DEPS_OK = False
    MISSING.append("openpyxl  (pip install openpyxl)")


# ══════════════════════════════════════════════════════════════════
#  2  CONSTANTES GLOBALES
#  ─────────────────────────────────────────────────────────────────
#  Todas las "tuning knobs" del sistema están agrupadas aquí para
#  facilitar ajustes sin tener que buscar en el código.
# ══════════════════════════════════════════════════════════════════

# ── URLs y base de datos ──────────────────────────────────────────

# URL de la página de búsqueda por RUC del portal SUNAT
URL_SUNAT = (
    "https://e-consultaruc.sunat.gob.pe/"
    "cl-ti-itmrconsruc/FrameCriterioBusquedaWeb.jsp"
)

# Sufijo que se añade al nombre del archivo Excel para el archivo
# SQLite de cache. Ej: "proveedores.xlsx" → "proveedores_sunat_v21.db"
DB_SUFFIX = "_sunat_v21.db"

# ── Paralelismo ───────────────────────────────────────────────────

# Número fijo de workers (instancias de Chromium).
# 2 es el máximo recomendado para no disparar el WAF de SUNAT.
NUM_WORKERS = 2

# ── Tiempos de pausa (todos en segundos) ─────────────────────────

# Pausa mínima/máxima entre etapas del mismo RUC
# (RUC→Deuda, Deuda→Trabajadores). Simula lectura humana.
PAUSA_ETAPA_MIN = 3.0
PAUSA_ETAPA_MAX = 6.0

# Con probabilidad PAUSA_LECTURA_PROB se añade una pausa larga extra
# a la pausa entre RUCs, simulando que el usuario "leyó" los datos.
PAUSA_LECTURA_PROB = 0.20   # 20 % de los RUCs tendrán pausa larga
PAUSA_LECTURA_MIN  = 8.0    # segundos mínimos de pausa larga
PAUSA_LECTURA_MAX  = 18.0   # segundos máximos de pausa larga

# ── Rate limiter global ───────────────────────────────────────────

# Intervalo mínimo (segundos) entre consultas "sensibles"
# (Deuda Coactiva y Trabajadores) considerando TODOS los workers.
# Evita que dos workers lancen peticiones simultáneas a esas URLs.
RATE_LIMIT_SEG = 4.0

# ── Timeouts y reintentos ─────────────────────────────────────────

# Tiempo máximo (segundos) para que Playwright espere una respuesta
TIMEOUT_CARGA = 28

# Veces que se reintenta un RUC antes de marcarlo como fallido
MAX_REINTENTOS = 3

# Backoff exponencial: 1.er reintento espera 8 s, 2.º 16 s, 3.º 32 s
ESPERA_REINTENTO_BASE = 8.0

# Cuántas veces puede reiniciarse Chromium por worker antes de rendirse
MAX_DRIVER_REINICIOS = 3

# ── Comportamiento WAF ────────────────────────────────────────────

# Pausa al detectar un bloqueo WAF individual (un solo worker)
ESPERA_WAF = 35.0

# Pausa cuando el umbral de workers bloqueados supera UMBRAL_WAF_IP;
# se asume que la IP está bloqueada globalmente.
PAUSA_IP_BLOQUEADA = 180.0

# Fracción de workers en estado WAF para activar la pausa global.
# 0.5 = si ≥ 50 % de los workers están bloqueados → pausa de IP.
UMBRAL_WAF_IP = 0.5

# ── ETA ───────────────────────────────────────────────────────────

# Cuántos tiempos recientes se promedian para calcular el ETA
VENTANA_ETA = 30

# ── Nombres de etapas (claves internas) ──────────────────────────

ETAPA_RUC          = "ruc_principal"
ETAPA_DEUDA        = "deuda_coactiva"
ETAPA_TRABAJADORES = "trabajadores"

# Clave especial en el dict de cache para marcar RUCs inválidos
FLAG_INVALIDO = "__invalido__"

# ── Mensajes reconocibles de SUNAT ───────────────────────────────

# Texto que aparece en la página cuando no hay deuda coactiva
MSG_SIN_DEUDA = "No se ha remitido deuda en cobranza coactiva"

# Texto que aparece cuando no hay declaraciones de trabajadores
MSG_SIN_TRABAJADORES = "No existen declaraciones presentadas"

# Fragmentos de texto en alerts de SUNAT que indican RUC inválido
PATRONES_RUC_INVALIDO = [
    "ingrese numero de ruc valido", "ingrese número de ruc válido",
    "ruc valido", "ruc válido", "numero de ruc", "número de ruc",
]

# ── Pool de User-Agents ───────────────────────────────────────────
# Solo Chrome/Firefox modernos en Windows/Mac para parecer tráfico real.
# Se seleccionan aleatoriamente al crear cada instancia de Chromium.
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 11.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]


# ══════════════════════════════════════════════════════════════════
#  3  RATE LIMITER GLOBAL
#  ─────────────────────────────────────────────────────────────────
#  Garantiza que las consultas "sensibles" (Deuda Coactiva y
#  Trabajadores) no se lancen a una frecuencia mayor que 1 cada
#  RATE_LIMIT_SEG segundos, sumando todos los workers.
#
#  Sin esto, dos workers podrían hacer peticiones casi simultáneas
#  a las mismas URLs internas de SUNAT y disparar el WAF.
# ══════════════════════════════════════════════════════════════════

class _RateLimiter:
    """
    Semáforo de frecuencia global (thread-safe).

    Garantiza un intervalo mínimo de `intervalo` segundos entre
    llamadas sucesivas a adquirir(), independientemente de desde
    qué hilo se llame.  El hilo que llama bloquea (sleep) hasta
    que pueda proceder sin violar el intervalo.
    """

    def __init__(self, intervalo: float):
        """
        Args:
            intervalo: Segundos mínimos entre llamadas permitidas.
        """
        self._intervalo = intervalo
        self._ultimo    = 0.0          # timestamp de la última llamada permitida
        self._lock      = threading.Lock()

    def adquirir(self):
        """
        Bloquea el hilo actual hasta que haya pasado `intervalo`
        segundos desde la última llamada permitida, luego registra
        este momento y regresa.
        """
        while True:
            with self._lock:
                ahora     = time.time()
                siguiente = self._ultimo + self._intervalo
                if ahora >= siguiente:
                    self._ultimo = ahora
                    return                    # se puede proceder
                espera = siguiente - ahora
            time.sleep(espera)               # espera fuera del lock


# Instancia global compartida por todos los workers.
# Se usa en consultar_deuda_coactiva() y consultar_trabajadores().
_rate_sensibles = _RateLimiter(RATE_LIMIT_SEG)


# ══════════════════════════════════════════════════════════════════
#  4  UTILIDADES PURAS
#  ─────────────────────────────────────────────────────────────────
#  Funciones helper sin efectos secundarios ni estado global.
#  Se definen antes que cualquier función de negocio para estar
#  disponibles en todo lo que sigue.
# ══════════════════════════════════════════════════════════════════

def _limpio(txt: str) -> str:
    """
    Normaliza una cadena de texto extraída del HTML de SUNAT.

    - Elimina espacios en blanco iniciales/finales.
    - Colapsa espacios/tabulaciones/saltos de línea internos a un solo espacio.
    - Devuelve "-" si el resultado está vacío o la entrada es None/falsy.

    Args:
        txt: Texto crudo extraído del DOM.

    Returns:
        Texto normalizado, o "-" si está vacío.
    """
    if not txt:
        return "-"
    txt = re.sub(r"\s+", " ", txt.strip())
    return txt or "-"


def _es_driver_vivo(driver_tuple) -> bool:
    """
    Comprueba si una instancia de Playwright/Chromium sigue activa.

    Intenta acceder al objeto browser y consultar is_connected().
    Si falla por cualquier razón, devuelve False para que el worker
    sepa que debe reiniciar el driver.

    Args:
        driver_tuple: Tupla (playwright, browser, page) o None.

    Returns:
        True si el browser responde, False en caso contrario.
    """
    try:
        if driver_tuple is None:
            return False
        _, browser, _ = driver_tuple
        return browser.is_connected()
    except Exception:
        return False


def _validar_ruc_local(ruc: str):
    """
    Validación sintáctica de un RUC antes de consultar SUNAT.

    Verifica que el RUC:
      - No esté vacío.
      - Contenga solo dígitos.
      - Tenga exactamente 11 dígitos.

    Esta validación evita viajes innecesarios a la red para RUCs
    claramente malformados (celdas con texto, fechas, etc.).

    Args:
        ruc: Valor tal como viene de la celda Excel.

    Returns:
        None si el RUC es válido, o un str con la descripción del error.
    """
    raw = str(ruc).strip()
    if not raw:
        return "RUC vacio"
    if not raw.isdigit():
        malos = sorted(set(c for c in raw if not c.isdigit()))
        desc  = ", ".join(f"'{c}'" for c in malos)
        return f"RUC con caracteres invalidos ({desc}): '{raw}'"
    if len(raw) != 11:
        return f"RUC debe tener 11 digitos, tiene {len(raw)}: '{raw}'"
    return None


def _pausa_entre_rucs() -> float:
    """
    Calcula la pausa aleatoria que el worker debe hacer al terminar
    un RUC completo y antes de empezar el siguiente.

    Base: 4–9 segundos (uniforme).
    Extra: Con probabilidad PAUSA_LECTURA_PROB (20 %) se añaden
           PAUSA_LECTURA_MIN–PAUSA_LECTURA_MAX segundos adicionales,
           simulando que el usuario se detuvo a leer los resultados.

    Returns:
        Segundos totales a dormir (float).
    """
    base = random.uniform(4.0, 9.0)
    if random.random() < PAUSA_LECTURA_PROB:
        base += random.uniform(PAUSA_LECTURA_MIN, PAUSA_LECTURA_MAX)
    return base


def _pausa_entre_etapas() -> float:
    """
    Calcula la pausa aleatoria entre las etapas de un mismo RUC
    (de RUC→Deuda o de Deuda→Trabajadores).

    Returns:
        Segundos a dormir entre etapas (float, uniforme en
        [PAUSA_ETAPA_MIN, PAUSA_ETAPA_MAX]).
    """
    return random.uniform(PAUSA_ETAPA_MIN, PAUSA_ETAPA_MAX)


# ══════════════════════════════════════════════════════════════════
#  5  PERSISTENCIA SQLITE (CACHE)
#  ─────────────────────────────────────────────────────────────────
#  Cada consulta exitosa (o fallida de forma definitiva) se persiste
#  en un archivo SQLite junto al Excel de entrada.  Al relanzar la
#  app con el mismo archivo, los RUCs ya procesados se cargan del
#  cache y no se vuelven a consultar en SUNAT.
#
#  Esquema de la tabla "resultados":
#    ruc        TEXT PRIMARY KEY  — 11 dígitos
#    datos_json TEXT              — dict serializado con las 3 etapas
#    ts         TEXT              — ISO timestamp del último guardado
# ══════════════════════════════════════════════════════════════════

def _db_path(ruta_excel: str) -> str:
    """
    Deriva la ruta del archivo SQLite a partir del Excel de entrada.

    Reemplaza la extensión del Excel por DB_SUFFIX.  Ejemplo:
    "/data/proveedores.xlsx" → "/data/proveedores_sunat_v21.db"

    Args:
        ruta_excel: Ruta absoluta o relativa del archivo Excel.

    Returns:
        Ruta del archivo SQLite correspondiente.
    """
    return os.path.splitext(ruta_excel)[0] + DB_SUFFIX


def db_inicializar(ruta_excel: str) -> sqlite3.Connection:
    """
    Abre (o crea) la base de datos SQLite para el Excel dado.

    Crea la tabla "resultados" si no existe.  La conexión se abre
    con check_same_thread=False porque múltiples workers la usan
    concurrentemente (protegida externamente con db_lock).

    Args:
        ruta_excel: Ruta del Excel; se usa para derivar la ruta del .db.

    Returns:
        Conexión sqlite3 lista para usar.
    """
    conn = sqlite3.connect(_db_path(ruta_excel), check_same_thread=False)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS resultados (
            ruc        TEXT PRIMARY KEY,
            datos_json TEXT NOT NULL,
            ts         TEXT NOT NULL
        )
    """)
    conn.commit()
    return conn


def db_guardar(conn: sqlite3.Connection, ruc: str, datos: dict):
    """
    Inserta o reemplaza el resultado de un RUC en la base de datos.

    El dict `datos` contiene las claves ETAPA_RUC, ETAPA_DEUDA y
    ETAPA_TRABAJADORES con los dicts parciales de cada etapa.

    Args:
        conn:  Conexión SQLite abierta.
        ruc:   Número de RUC (11 dígitos como str).
        datos: Dict con los resultados crudos de las 3 etapas.
    """
    conn.execute(
        "INSERT OR REPLACE INTO resultados (ruc, datos_json, ts) VALUES (?, ?, ?)",
        (str(ruc), json.dumps(datos, ensure_ascii=False), datetime.now().isoformat())
    )
    conn.commit()


def db_cargar_procesados(conn: sqlite3.Connection) -> dict:
    """
    Carga todos los RUCs ya procesados desde la base de datos.

    Returns:
        Dict {ruc_str: datos_dict} con todos los registros del cache.
    """
    cur = conn.execute("SELECT ruc, datos_json FROM resultados")
    return {row[0]: json.loads(row[1]) for row in cur.fetchall()}


def db_limpiar(conn: sqlite3.Connection):
    """
    Borra todos los registros de la tabla de resultados.

    Se invoca cuando el usuario desactiva el cache antes de iniciar,
    para forzar una consulta fresca de todos los RUCs.

    Args:
        conn: Conexión SQLite abierta.
    """
    conn.execute("DELETE FROM resultados")
    conn.commit()


# ══════════════════════════════════════════════════════════════════
#  6  PLAYWRIGHT — GESTIÓN DEL DRIVER (CHROMIUM)
#  ─────────────────────────────────────────────────────────────────
#  Estas funciones encapsulan la creación y destrucción de instancias
#  de Chromium.  Cada worker mantiene su propio driver_tuple.
#
#  driver_tuple = (playwright_instance, browser, page)
#
#  Estrategia de evasión WAF:
#    - Resolución de viewport realista (1366×768)
#    - User-agent rotado del pool USER_AGENTS
#    - Locale español Perú
#    - Imágenes bloqueadas (reduce tráfico sin afectar scraping)
#    - Sin trucos navigator.* (frágiles y detectables)
#    - La evasión real viene de las pausas humanas (4)
# ══════════════════════════════════════════════════════════════════

def iniciar_driver(user_agent: str = None):
    """
    Crea y configura una nueva instancia de Playwright + Chromium.

    Configura:
    - Viewport 1366×768 (la resolución de escritorio más común).
    - User-agent tomado del pool o el proporcionado.
    - Locale es-PE para parecer un usuario peruano.
    - Bloqueo de imágenes/fuentes para acelerar las cargas.
    - Timeout global de TIMEOUT_CARGA segundos.

    Args:
        user_agent: UA string personalizado. Si es None, se elige
                    aleatoriamente del pool USER_AGENTS.

    Returns:
        Tupla (playwright, browser, page) lista para usar.
    """
    ua      = user_agent or random.choice(USER_AGENTS)
    pw      = sync_playwright().start()
    browser = pw.chromium.launch(
        headless=True,
        args=[
            "--no-sandbox",
            "--disable-dev-shm-usage",
            "--disable-gpu",
            "--ignore-certificate-errors",
        ],
    )
    context = browser.new_context(
        viewport={"width": 1366, "height": 768},
        user_agent=ua,
        locale="es-PE",
        java_script_enabled=True,
        ignore_https_errors=True,
    )
    # Bloquear recursos visuales: reduce tráfico de red
    # sin afectar el DOM que se parsea para extraer datos.
    context.route(
        "**/*.{png,jpg,jpeg,gif,svg,webp,ico,woff,woff2}",
        lambda r: r.abort()
    )
    page = context.new_page()
    page.set_default_timeout(TIMEOUT_CARGA * 1000)
    return (pw, browser, page)


def cerrar_driver_seguro(driver_tuple):
    """
    Cierra todos los componentes de un driver_tuple de forma segura.

    Intenta cerrar page → browser → playwright en ese orden,
    ignorando cualquier excepción (el browser puede ya estar muerto).

    Args:
        driver_tuple: Tupla (playwright, browser, page) o None.
    """
    if not driver_tuple:
        return
    pw, browser, page = driver_tuple
    for obj, method in [(page, "close"), (browser, "close"), (pw, "stop")]:
        try:
            getattr(obj, method)()
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════
#  7  EXCEPCIONES PERSONALIZADAS
#  ─────────────────────────────────────────────────────────────────
#  Excepciones propias del dominio para controlar el flujo en los
#  workers sin mezclarlas con errores genéricos de Playwright.
# ══════════════════════════════════════════════════════════════════

class WAFBlockException(Exception):
    """
    Se lanza cuando SUNAT devuelve una página de bloqueo WAF
    ("The requested URL was rejected" / "support ID").

    El worker la captura para aplicar la pausa correspondiente y
    reintentar o marcar el RUC como fallido.
    """


class RUCInvalidoException(Exception):
    """
    Se lanza cuando SUNAT muestra un alert indicando que el RUC
    buscado no existe o es inválido según su base de datos.

    A diferencia de un error transitorio, este RUC no debe
    reintentarse; se marca directamente como inválido.
    """


# ══════════════════════════════════════════════════════════════════
#  8  CAPA DE SCRAPING — CONSULTAS SUNAT
#  ─────────────────────────────────────────────────────────────────
#  Tres funciones independientes, cada una encargada de una "pantalla"
#  del portal SUNAT.  Se llaman siempre en este orden:
#
#    1. consultar_ruc_principal()   ← navega a URL_SUNAT + submit form
#    2. consultar_deuda_coactiva()  ← POST interno (accion=getInfoDC)
#    3. consultar_trabajadores()    ← POST interno (accion=getCantTrab)
#
#  Nota sobre _submit_form():
#    Las pantallas 2 y 3 no tienen un link directo; se navega a ellas
#    creando y enviando dinámicamente un <form> POST desde JavaScript,
#    replicando lo que hace el portal manualmente.
# ══════════════════════════════════════════════════════════════════

def _submit_form(page, accion: str, nro_ruc: str, des_ruc: str):
    """
    Simula el envío del formulario de navegación interno de SUNAT.

    El portal usa formularios POST para moverse entre secciones
    (deuda, trabajadores).  Esta función inyecta y envía dicho
    formulario desde JavaScript, evitando tener que encontrar y
    hacer clic en botones de la UI.

    Args:
        page:    Objeto Page de Playwright (la pestaña activa).
        accion:  Valor del campo "accion" del formulario.
                 "getInfoDC"    → Deuda Coactiva
                 "getCantTrab"  → Trabajadores
        nro_ruc: RUC a consultar (11 dígitos).
        des_ruc: Nombre/razón social del contribuyente (obtenido en
                 la etapa RUC; SUNAT lo usa para pre-rellenar la página).
    """
    page.evaluate(
        """([accion, nroRuc, desRuc]) => {
            var form = document.createElement('form');
            form.method = 'POST';
            form.action = '/cl-ti-itmrconsruc/jcrS00Alias';
            var fields = {
                accion: accion, contexto: 'ti-it',
                modo: '1', nroRuc: nroRuc, desRuc: desRuc
            };
            for (var n in fields) {
                var i = document.createElement('input');
                i.type = 'hidden'; i.name = n; i.value = fields[n];
                form.appendChild(i);
            }
            document.body.appendChild(form);
            form.submit();
        }""",
        [accion, nro_ruc, des_ruc],
    )


def consultar_ruc_principal(driver_tuple, ruc: str) -> dict:
    """
    Etapa 1: Consulta los datos generales del contribuyente en SUNAT.

    Navega a URL_SUNAT, ingresa el RUC en el campo de búsqueda,
    hace clic en "Aceptar" y parsea el panel de resultados.

    Datos extraídos:
        - Tipo de contribuyente
        - Nombre comercial
        - Fecha de inscripción / Fecha de inicio de actividades
        - Estado y condición del contribuyente
        - Domicilio fiscal
        - Actividades económicas (principal y todas)

    El nombre/razón social se guarda en "_nombre_ruc" (clave interna)
    porque SUNAT lo muestra junto al número en el encabezado.

    Manejo de errores:
        - Alert con patrón de RUC inválido → RUCInvalidoException
        - Respuesta WAF → WAFBlockException
        - Timeout / error de parseo → se devuelve dict con "error_etapa"

    Args:
        driver_tuple: Tupla (playwright, browser, page).
        ruc:          Número de RUC a consultar (str, 11 dígitos).

    Returns:
        Dict con los campos de la etapa RUC.  Si hay error,
        "error_etapa" contendrá la descripción.
    """
    _, _, page = driver_tuple
    TO = TIMEOUT_CARGA * 1000   # Playwright usa milisegundos

    # Estructura base con valores por defecto "-" para todos los campos
    base = {
        "RUC":                           str(ruc),
        "Tipo_Contribuyente":            "-",
        "Nombre_Comercial":              "-",
        "Fecha_Inscripcion":             "-",
        "Fecha_Inicio_Actividades":      "-",
        "Estado_Contribuyente":          "-",
        "Condicion_Contribuyente":       "-",
        "Domicilio_Fiscal":              "-",
        "Actividad_Economica_Principal": "-",
        "Actividades_Economicas":        "-",
        "error_etapa":                   "",
    }

    # Capturamos alerts del navegador para detectar RUC inválido
    _alerts = []
    def _on_dialog(dialog):
        _alerts.append(dialog.message)
        dialog.accept()
    page.on("dialog", _on_dialog)

    # ── Navegación y búsqueda ─────────────────────────────────────
    try:
        page.goto(URL_SUNAT, wait_until="domcontentloaded", timeout=TO)
        page.wait_for_selector("#txtRuc", timeout=TO)

        # Pequeña pausa antes de escribir, como haría un humano
        time.sleep(random.uniform(0.4, 1.0))
        page.fill("#txtRuc", str(ruc).strip())
        time.sleep(random.uniform(0.3, 0.8))
        page.click("#btnAceptar")
        time.sleep(0.8)
        page.remove_listener("dialog", _on_dialog)
    except PlaywrightTimeoutError as e:
        base["error_etapa"] = f"Timeout cargando busqueda: {str(e)[:80]}"
        return base

    # ── Verificar alert de RUC inválido ──────────────────────────
    if _alerts:
        txt = _alerts[0]
        if any(p in txt.lower() for p in PATRONES_RUC_INVALIDO):
            raise RUCInvalidoException(f"RUC invalido SUNAT: {txt.strip()}")
        base["error_etapa"] = f"Alert: {txt.strip()[:100]}"
        return base

    # ── Esperar resultados o error ────────────────────────────────
    try:
        page.wait_for_selector(
            ".panel.panel-primary, .alert, .mensajeError",
            timeout=TO
        )
    except PlaywrightTimeoutError as e:
        base["error_etapa"] = f"Timeout esperando resultado: {str(e)[:80]}"
        return base

    # ── Detectar bloqueo WAF ──────────────────────────────────────
    src = page.content() or ""
    if "The requested URL was rejected" in src or "support ID" in src:
        raise WAFBlockException(f"WAF SUNAT RUC {ruc}")

    # ── Verificar mensajes de error en la página ──────────────────
    for sel in [".alert", ".mensajeError", "#mensError"]:
        el = page.query_selector(sel)
        if el:
            txt = (el.inner_text() or "").strip()
            if txt:
                base["error_etapa"] = _limpio(txt)
                return base

    # ── Parsear panel de resultados ───────────────────────────────
    try:
        items = page.query_selector_all(
            ".panel.panel-primary .list-group .list-group-item"
        )
        for item in items:
            headings = item.query_selector_all("h4.list-group-item-heading")
            if not headings:
                continue
            label = _limpio(headings[0].inner_text()).upper()

            # Valor: primero busca <p>, si no usa segundo <h4>
            val_el = item.query_selector("p.list-group-item-text")
            if val_el:
                valor = _limpio(val_el.inner_text())
            elif len(headings) >= 2:
                valor = _limpio(headings[1].inner_text())
            else:
                valor = "-"

            # Mapeo label → campo de base
            if "NÚMERO DE RUC" in label or "NUMERO DE RUC" in label:
                # Formato: "20123456789 - RAZÓN SOCIAL S.A."
                m = re.match(r"^\d{11}\s*[-–]\s*(.+)$", valor)
                if m:
                    base["_nombre_ruc"] = _limpio(m.group(1))

            elif "TIPO CONTRIBUYENTE" in label:
                base["Tipo_Contribuyente"] = valor

            elif "NOMBRE COMERCIAL" in label:
                base["Nombre_Comercial"] = valor

            elif "FECHA DE INSCRIPCIÓN" in label or "FECHA DE INSCRIPCION" in label:
                # Este bloque puede contener dos fechas (inscripción + inicio)
                all_h4 = item.query_selector_all("h4.list-group-item-heading")
                all_p  = item.query_selector_all("p.list-group-item-text")
                for i_h, h in enumerate(all_h4):
                    lbl2 = _limpio(h.inner_text()).upper()
                    val2 = _limpio(all_p[i_h].inner_text()) if i_h < len(all_p) else "-"
                    if "FECHA DE INSCRIPCIÓN" in lbl2 or "FECHA DE INSCRIPCION" in lbl2:
                        base["Fecha_Inscripcion"] = val2
                    elif "FECHA DE INICIO DE ACTIVIDADES" in lbl2:
                        base["Fecha_Inicio_Actividades"] = val2

            elif "FECHA DE INICIO DE ACTIVIDADES" in label:
                base["Fecha_Inicio_Actividades"] = valor

            elif "ESTADO DEL CONTRIBUYENTE" in label:
                base["Estado_Contribuyente"] = valor

            elif ("CONDICIÓN DEL CONTRIBUYENTE" in label
                  or "CONDICION DEL CONTRIBUYENTE" in label):
                base["Condicion_Contribuyente"] = valor

            elif "DOMICILIO FISCAL" in label:
                base["Domicilio_Fiscal"] = valor

            elif "ACTIVIDAD" in label and ("ECONÓMICA" in label or "ECONOMICA" in label):
                # Las actividades están en celdas de una tabla interna
                tds = item.query_selector_all("table td")
                actividades = [
                    _limpio(td.inner_text()) for td in tds
                    if _limpio(td.inner_text()) != "-"
                ]
                if actividades:
                    base["Actividades_Economicas"] = " | ".join(actividades)
                    for act in actividades:
                        if act.upper().startswith("PRINCIPAL"):
                            m = re.search(r"Principal\s*-\s*(.+)", act, re.IGNORECASE)
                            base["Actividad_Economica_Principal"] = (
                                _limpio(m.group(1)) if m else act
                            )
                            break

    except Exception as e:
        base["error_etapa"] = f"Error parseando datos RUC: {str(e)[:100]}"
        return base

    base.setdefault("_nombre_ruc", "-")
    return base


def consultar_deuda_coactiva(driver_tuple, ruc: str, des_ruc: str) -> dict:
    """
    Etapa 2: Consulta si el contribuyente tiene deuda en cobranza coactiva.

    Primero adquiere el rate limiter global para no saturar SUNAT,
    luego navega a la sección de deuda usando _submit_form().

    Extrae la última fila de la tabla de deuda (la más reciente):
        - Monto de la deuda (S/)
        - Período tributario
        - Fecha de inicio de cobranza
        - Entidad asociada

    También captura la fecha de actualización del sistema SUNAT.

    Si SUNAT muestra MSG_SIN_DEUDA → Deuda_Tiene = "NO"
    Si hay tabla con filas         → Deuda_Tiene = "SI" + detalle
    Si no se puede determinar      → Deuda_Tiene = "DESCONOCIDO"

    Args:
        driver_tuple: Tupla (playwright, browser, page).
        ruc:          RUC a consultar.
        des_ruc:      Nombre/razón social (necesario para el POST).

    Returns:
        Dict con campos Deuda_*.  Si hay error, "error_etapa" lo describe.

    Raises:
        WAFBlockException: Si SUNAT devuelve página de bloqueo.
    """
    _rate_sensibles.adquirir()    # ← espera su turno en el rate limiter global
    _, _, page = driver_tuple
    TO = TIMEOUT_CARGA * 1000

    base = {
        "Deuda_Tiene":                 "-",
        "Deuda_Monto":                 "-",
        "Deuda_Periodo":               "-",
        "Deuda_Fecha_Inicio_Cobranza": "-",
        "Deuda_Entidad":               "-",
        "Deuda_Fecha_Actualizacion":   "-",
        "error_etapa":                 "",
    }

    # ── Navegar a la sección de deuda ─────────────────────────────
    try:
        _submit_form(page, "getInfoDC", str(ruc).strip(), des_ruc)
    except Exception as e:
        base["error_etapa"] = f"Error enviando formulario deuda: {str(e)[:80]}"
        return base

    # ── Esperar a que cargue la tabla O el mensaje "sin deuda" ────
    try:
        page.wait_for_function(
            "(msg) => {"
            "  const p = document.querySelector('.panel.panel-primary');"
            "  if (!p) return false;"
            "  if (p.querySelectorAll('table tbody tr').length > 0) return true;"
            "  return Array.from(document.querySelectorAll('.list-group-item,.col-sm-12'))"
            "    .some(el => el.textContent.includes(msg));"
            "}",
            arg=MSG_SIN_DEUDA,
            timeout=TO,
        )
    except PlaywrightTimeoutError:
        pass    # Se continúa; se analizará el contenido disponible

    # ── Detectar bloqueo WAF ──────────────────────────────────────
    src = page.content() or ""
    if "The requested URL was rejected" in src or "support ID" in src:
        raise WAFBlockException(f"WAF SUNAT deuda RUC {ruc}")

    # ── Capturar fecha de actualización del sistema ───────────────
    fecha_act = "-"
    try:
        for h4 in page.query_selector_all(".list-group-item h4"):
            m = re.search(
                r"actualizada al\s+(\d{2}/\d{2}/\d{4})",
                h4.inner_text() or ""
            )
            if m:
                fecha_act = m.group(1)
                break
    except Exception:
        pass
    base["Deuda_Fecha_Actualizacion"] = fecha_act

    # ── Parsear tabla de deuda ────────────────────────────────────
    try:
        tabla = page.query_selector(".panel.panel-primary table")
        if tabla:
            filas = tabla.query_selector_all("tbody tr")
            if filas:
                # Solo tomamos la última fila (deuda más reciente)
                celdas = filas[-1].query_selector_all("td")
                if len(celdas) >= 4:
                    base["Deuda_Tiene"]                = "SI"
                    base["Deuda_Monto"]                = _limpio(celdas[0].inner_text())
                    base["Deuda_Periodo"]              = _limpio(celdas[1].inner_text())
                    base["Deuda_Fecha_Inicio_Cobranza"] = _limpio(celdas[2].inner_text())
                    base["Deuda_Entidad"]              = _limpio(celdas[3].inner_text())
                    return base
    except Exception as e:
        base["error_etapa"] = f"Error leyendo tabla deuda: {str(e)[:80]}"
        return base

    # ── Verificar mensaje "sin deuda" ─────────────────────────────
    sin_deuda = any(
        MSG_SIN_DEUDA in (el.inner_text() or "")
        for el in page.query_selector_all(".list-group-item, .col-sm-12")
    )
    if sin_deuda:
        base["Deuda_Tiene"] = "NO"
    else:
        base["Deuda_Tiene"]  = "DESCONOCIDO"
        base["error_etapa"]  = "No se pudo determinar estado de deuda"
    return base


def consultar_trabajadores(driver_tuple, ruc: str, des_ruc: str) -> dict:
    """
    Etapa 3: Consulta el histórico de declaraciones de trabajadores.

    Adquiere el rate limiter global y navega a la sección de
    trabajadores usando _submit_form().

    Extrae la última fila del histórico (período más reciente):
        - Período
        - Número de trabajadores
        - Número de pensionistas
        - Número de prestadores de servicio

    Si SUNAT muestra MSG_SIN_TRABAJADORES → Trab_Tiene = "NO"
    Si hay tabla con filas               → Trab_Tiene = "SI" + detalle
    Si no se puede determinar            → Trab_Tiene = "DESCONOCIDO"

    Args:
        driver_tuple: Tupla (playwright, browser, page).
        ruc:          RUC a consultar.
        des_ruc:      Nombre/razón social (necesario para el POST).

    Returns:
        Dict con campos Trab_*.  Si hay error, "error_etapa" lo describe.

    Raises:
        WAFBlockException: Si SUNAT devuelve página de bloqueo.
    """
    _rate_sensibles.adquirir()    # ← espera su turno en el rate limiter global
    _, _, page = driver_tuple
    TO = TIMEOUT_CARGA * 1000

    base = {
        "Trab_Tiene":            "-",
        "Trab_Periodo":          "-",
        "Trab_Num_Trabajadores": "-",
        "Trab_Num_Pensionistas": "-",
        "Trab_Num_Prestadores":  "-",
        "error_etapa":           "",
    }

    # ── Navegar a la sección de trabajadores ──────────────────────
    try:
        _submit_form(page, "getCantTrab", str(ruc).strip(), des_ruc)
    except Exception as e:
        base["error_etapa"] = f"Error enviando formulario trabajadores: {str(e)[:80]}"
        return base

    # ── Esperar panel de resultados ───────────────────────────────
    try:
        page.wait_for_selector(".panel.panel-primary", timeout=TO)
    except PlaywrightTimeoutError as e:
        base["error_etapa"] = f"Timeout cargando trabajadores: {str(e)[:60]}"
        return base

    # ── Detectar bloqueo WAF ──────────────────────────────────────
    src = page.content() or ""
    if "The requested URL was rejected" in src or "support ID" in src:
        raise WAFBlockException(f"WAF SUNAT trabajadores RUC {ruc}")

    # ── Verificar mensaje "sin declaraciones" ─────────────────────
    sin_decl = any(
        MSG_SIN_TRABAJADORES in (el.inner_text() or "")
        for el in page.query_selector_all("h5, .panel-primary h5, .col-md-12 h5")
    )
    if sin_decl:
        base["Trab_Tiene"] = "NO"
        return base

    # ── Parsear tabla de trabajadores ─────────────────────────────
    try:
        tabla = page.query_selector(".panel.panel-primary table")
        if tabla:
            filas = tabla.query_selector_all("tbody tr")
            if filas:
                # Solo tomamos la última fila (período más reciente)
                celdas = filas[-1].query_selector_all("td")
                if len(celdas) >= 4:
                    base["Trab_Tiene"]            = "SI"
                    base["Trab_Periodo"]          = _limpio(celdas[0].inner_text())
                    base["Trab_Num_Trabajadores"] = _limpio(celdas[1].inner_text())
                    base["Trab_Num_Pensionistas"] = _limpio(celdas[2].inner_text())
                    base["Trab_Num_Prestadores"]  = _limpio(celdas[3].inner_text())
                    return base
    except Exception as e:
        base["error_etapa"] = f"Error leyendo tabla trabajadores: {str(e)[:80]}"
        return base

    base["Trab_Tiene"]  = "DESCONOCIDO"
    base["error_etapa"] = "No se pudo determinar datos de trabajadores"
    return base


# ══════════════════════════════════════════════════════════════════
#  9  ORQUESTACIÓN DE CONSULTA COMPLETA
#  ─────────────────────────────────────────────────────────────────
#  consultar_ruc_completo() coordina las 3 etapas de scraping para
#  un RUC dado, reutilizando resultados previos del cache cuando
#  alguna etapa ya fue completada exitosamente.
#
#  Las funciones _armar_resultado_final() y _armar_fila_invalida()
#  construyen el dict "plano" final que se usa en la UI y en el Excel.
#
#  _etapas_pendientes() se usa en el worker para saber si el cache
#  ya cubre todas las etapas de un RUC antes de consultar SUNAT.
# ══════════════════════════════════════════════════════════════════

def consultar_ruc_completo(driver_tuple, ruc: str, cache_previo: dict = None) -> dict:
    """
    Orquesta las 3 etapas de consulta para un RUC, con cache granular.

    Lógica de cache:
        - Si una etapa ya está en cache_previo y no tiene error,
          se reutiliza directamente (no se vuelve a consultar SUNAT).
        - Si la etapa RUC falla, las etapas de Deuda y Trabajadores
          se marcan como "saltadas" y no se intentan.

    Pausas entre etapas:
        Se duerme _pausa_entre_etapas() entre cada etapa para
        simular comportamiento humano.

    Args:
        driver_tuple:  Tupla (playwright, browser, page).
        ruc:           RUC a consultar (str, 11 dígitos).
        cache_previo:  Dict {ETAPA_*: dict_etapa} con resultados
                       previos (puede ser None o vacío).

    Returns:
        Dict plano con todos los campos (ver _armar_resultado_final).
        El campo "completado" es True si las 3 etapas salieron sin error.
    """
    cache     = cache_previo or {}
    resultado = {
        ETAPA_RUC:          cache.get(ETAPA_RUC),
        ETAPA_DEUDA:        cache.get(ETAPA_DEUDA),
        ETAPA_TRABAJADORES: cache.get(ETAPA_TRABAJADORES),
    }

    def _ok(etapa):
        """Retorna True si la etapa tiene datos y no tiene error."""
        d = resultado.get(etapa)
        return d is not None and not d.get("error_etapa", "")

    # ── Etapa 1: RUC Principal ────────────────────────────────────
    if not _ok(ETAPA_RUC):
        datos = consultar_ruc_principal(driver_tuple, ruc)
        resultado[ETAPA_RUC] = datos
        if datos.get("error_etapa"):
            # Si RUC falla, no tiene sentido intentar las otras etapas
            resultado[ETAPA_DEUDA] = (
                resultado[ETAPA_DEUDA]
                or {"error_etapa": "Saltado por fallo en RUC principal"}
            )
            resultado[ETAPA_TRABAJADORES] = (
                resultado[ETAPA_TRABAJADORES]
                or {"error_etapa": "Saltado por fallo en RUC principal"}
            )
            return _armar_resultado_final(ruc, resultado)

    # El nombre se necesita para los POSTs de las siguientes etapas
    des_ruc = resultado[ETAPA_RUC].get("_nombre_ruc", "-")

    time.sleep(_pausa_entre_etapas())   # pausa humana entre etapas

    # ── Etapa 2: Deuda Coactiva ───────────────────────────────────
    if not _ok(ETAPA_DEUDA):
        resultado[ETAPA_DEUDA] = consultar_deuda_coactiva(driver_tuple, ruc, des_ruc)

    time.sleep(_pausa_entre_etapas())   # pausa humana entre etapas

    # ── Etapa 3: Trabajadores ─────────────────────────────────────
    if not _ok(ETAPA_TRABAJADORES):
        resultado[ETAPA_TRABAJADORES] = consultar_trabajadores(driver_tuple, ruc, des_ruc)

    return _armar_resultado_final(ruc, resultado)


def _armar_resultado_final(ruc: str, resultado: dict) -> dict:
    """
    Construye el dict "plano" final a partir de los dicts de las 3 etapas.

    Aplana los tres sub-dicts (RUC, Deuda, Trabajadores) en un único
    dict con todos los campos necesarios para la UI y el Excel.

    También agrega campos de control:
        - "Error":          String con todos los errores de etapas fallidas.
        - "Etapas_Fallidas": Lista de nombres de etapas con error.
        - "completado":     True si ninguna etapa falló.
        - "es_invalido":    True si el RUC fue marcado como inválido.
        - "_raw":           El dict original de resultados por etapa
                            (se usa para guardar en SQLite).

    Args:
        ruc:       RUC consultado.
        resultado: Dict {ETAPA_*: dict_etapa} con resultados de las 3 etapas.

    Returns:
        Dict plano con todos los campos de la fila.
    """
    ruc_d  = resultado.get(ETAPA_RUC)          or {}
    deu_d  = resultado.get(ETAPA_DEUDA)         or {}
    trab_d = resultado.get(ETAPA_TRABAJADORES)  or {}

    # Recolectar errores de cada etapa
    errores = {}
    if ruc_d.get("error_etapa"):   errores[ETAPA_RUC]          = ruc_d["error_etapa"]
    if deu_d.get("error_etapa"):   errores[ETAPA_DEUDA]        = deu_d["error_etapa"]
    if trab_d.get("error_etapa"):  errores[ETAPA_TRABAJADORES] = trab_d["error_etapa"]

    return {
        "RUC":                           str(ruc),
        "Nombre":                        ruc_d.get("_nombre_ruc", "-"),
        "Tipo_Contribuyente":            ruc_d.get("Tipo_Contribuyente", "-"),
        "Nombre_Comercial":              ruc_d.get("Nombre_Comercial", "-"),
        "Fecha_Inscripcion":             ruc_d.get("Fecha_Inscripcion", "-"),
        "Fecha_Inicio_Actividades":      ruc_d.get("Fecha_Inicio_Actividades", "-"),
        "Estado_Contribuyente":          ruc_d.get("Estado_Contribuyente", "-"),
        "Condicion_Contribuyente":       ruc_d.get("Condicion_Contribuyente", "-"),
        "Domicilio_Fiscal":              ruc_d.get("Domicilio_Fiscal", "-"),
        "Actividad_Economica_Principal": ruc_d.get("Actividad_Economica_Principal", "-"),
        "Actividades_Economicas":        ruc_d.get("Actividades_Economicas", "-"),
        "Deuda_Tiene":                   deu_d.get("Deuda_Tiene", "-"),
        "Deuda_Monto":                   deu_d.get("Deuda_Monto", "-"),
        "Deuda_Periodo":                 deu_d.get("Deuda_Periodo", "-"),
        "Deuda_Fecha_Inicio_Cobranza":   deu_d.get("Deuda_Fecha_Inicio_Cobranza", "-"),
        "Deuda_Entidad":                 deu_d.get("Deuda_Entidad", "-"),
        "Trab_Tiene":                    trab_d.get("Trab_Tiene", "-"),
        "Trab_Periodo":                  trab_d.get("Trab_Periodo", "-"),
        "Trab_Num_Trabajadores":         trab_d.get("Trab_Num_Trabajadores", "-"),
        "Trab_Num_Pensionistas":         trab_d.get("Trab_Num_Pensionistas", "-"),
        "Trab_Num_Prestadores":          trab_d.get("Trab_Num_Prestadores", "-"),
        "Error":          " | ".join(f"[{e}] {m}" for e, m in errores.items()),
        "Etapas_Fallidas": list(errores.keys()),
        "completado":     len(errores) == 0,
        "es_invalido":    bool(resultado.get(FLAG_INVALIDO)),
        "_raw":           resultado,
    }


def _armar_fila_invalida(ruc: str, motivo: str) -> dict:
    """
    Construye una fila de resultado para un RUC que no se pudo consultar.

    Se usa en dos situaciones:
        1. Validación local fallida (_validar_ruc_local devolvió error).
        2. SUNAT confirmó que el RUC no existe (RUCInvalidoException).
        3. El worker agotó los reintentos (fallido definitivo).

    En todos los casos, todos los campos quedan en "-" excepto
    "Estado_Contribuyente" = "INVALIDO" y "Error" con el motivo.

    Args:
        ruc:    RUC problemático (puede tener formato incorrecto).
        motivo: Descripción textual del motivo de invalidez.

    Returns:
        Dict plano compatible con la UI y la exportación Excel.
    """
    return {
        "RUC":                           str(ruc),
        "Nombre":                        "-",
        "Tipo_Contribuyente":            "-",
        "Nombre_Comercial":              "-",
        "Fecha_Inscripcion":             "-",
        "Fecha_Inicio_Actividades":      "-",
        "Estado_Contribuyente":          "INVALIDO",
        "Condicion_Contribuyente":       "-",
        "Domicilio_Fiscal":              "-",
        "Actividad_Economica_Principal": "-",
        "Actividades_Economicas":        "-",
        "Deuda_Tiene":                   "-",
        "Deuda_Monto":                   "-",
        "Deuda_Periodo":                 "-",
        "Deuda_Fecha_Inicio_Cobranza":   "-",
        "Deuda_Entidad":                 "-",
        "Deuda_Fecha_Actualizacion":     "-",
        "Trab_Tiene":                    "-",
        "Trab_Periodo":                  "-",
        "Trab_Num_Trabajadores":         "-",
        "Trab_Num_Pensionistas":         "-",
        "Trab_Num_Prestadores":          "-",
        "Error":           motivo,
        "Etapas_Fallidas": [ETAPA_RUC],
        "completado":      False,
        "es_invalido":     True,
        "_raw":            {},
    }


def _etapas_pendientes(cache: dict) -> list:
    """
    Determina qué etapas de un RUC aún necesitan consultarse.

    Revisa el cache de un RUC y devuelve la lista de etapas que
    no tienen resultado o que tienen un error guardado.

    Se usa en el worker para decidir si vale la pena lanzar una
    consulta completa o si el RUC ya está cubierto por el cache.

    Args:
        cache: Dict {ETAPA_*: dict_etapa} o dict vacío / None.

    Returns:
        Lista de constantes ETAPA_* que faltan (puede estar vacía).
    """
    if not cache:
        return [ETAPA_RUC, ETAPA_DEUDA, ETAPA_TRABAJADORES]
    return [
        e for e in [ETAPA_RUC, ETAPA_DEUDA, ETAPA_TRABAJADORES]
        if not cache.get(e) or cache[e].get("error_etapa")
    ]


# ══════════════════════════════════════════════════════════════════
#  10  WORKER
#  ─────────────────────────────────────────────────────────────────
#  Función que corre en un hilo separado por cada worker.
#  Consume RUCs de ruc_queue y produce resultados en result_queue.
#
#  Ciclo de vida de un worker:
#    1. Inicia Chromium (iniciar_driver)
#    2. Loop: toma RUC de la cola
#       a. Validación local → si inválido, guarda y pasa al siguiente
#       b. Hasta MAX_REINTENTOS veces:
#            - consultar_ruc_completo()
#            - Si WAFBlockException → pausa y reintenta
#            - Si RUCInvalidoException → marca inválido, siguiente RUC
#            - Si PlaywrightError → reinicia Chromium si necesario
#       c. Guarda en SQLite, pone en result_queue
#       d. Duerme _pausa_entre_rucs()
#    3. Al salir del loop, cierra Chromium
#
#  Manejo de WAF:
#    - waf_counter es un int compartido entre todos los workers.
#    - Si activos/total >= UMBRAL_WAF_IP → pausa larga (la IP está bloqueada).
#    - Si < umbral → pausa individual (solo este worker fue detectado).
# ══════════════════════════════════════════════════════════════════

def worker_procesar_rucs(
    worker_id,
    ruc_queue,
    cancelar_flag,
    log_queue,
    result_queue,
    fallidos_queue,
    waf_counter,
    waf_counter_lock,
    total_workers,
    db_conn,
    db_lock,
    user_agent=None,
):
    """
    Función principal de cada hilo worker.

    Consume RUCs de ruc_queue hasta que la cola esté vacía o
    cancelar_flag() devuelva True.

    Args:
        worker_id:        Identificador numérico del worker (1, 2, …).
        ruc_queue:        Queue de tuplas (posicion_global, ruc, cache_previo).
        cancelar_flag:    Callable que devuelve True si se pidió cancelar.
        log_queue:        Queue para enviar mensajes de log a la UI.
                          Cada item es (tag, mensaje_str).
        result_queue:     Queue para enviar filas de resultado a la UI.
                          Cada item es {"pos": int, "fila": dict}.
        fallidos_queue:   Queue para RUCs que agotaron reintentos.
                          Cada item es (posicion, ruc, motivo).
        waf_counter:      Lista de un elemento [int] (contador de workers en WAF).
                          Se usa lista para mutabilidad compartida entre hilos.
        waf_counter_lock: Lock que protege el acceso a waf_counter.
        total_workers:    Total de workers activos (para calcular el umbral WAF).
        db_conn:          Conexión SQLite compartida para persistir resultados.
        db_lock:          Lock que serializa el acceso a db_conn.
        user_agent:       User-agent string para este worker (o None → aleatorio).
    """
    driver           = None
    driver_reinicios = 0    # contador de reinicios de Chromium para este worker

    def log(tag, msg):
        """Encola un mensaje de log con el prefijo del worker."""
        log_queue.put((tag, f"[W{worker_id:02d}] {msg}"))

    def iniciar_o_reiniciar(motivo=""):
        """
        Cierra el driver actual (si existe) y crea uno nuevo.

        Returns:
            True si el nuevo driver se inició correctamente, False si falló.
        """
        nonlocal driver, driver_reinicios
        cerrar_driver_seguro(driver)
        if motivo:
            log("warn", f"Reiniciando Chromium ({motivo})...")
        try:
            driver = iniciar_driver(user_agent)
            if motivo:
                log("info", "Chromium reiniciado.")
            return True
        except Exception as e:
            log("error", f"No se pudo iniciar Chromium: {e}")
            return False

    # ── Arranque del worker ───────────────────────────────────────
    if not iniciar_o_reiniciar():
        # Si no se puede iniciar Chromium, marcar todos los RUCs como fallidos
        while True:
            try:
                pos, ruc, cache = ruc_queue.get_nowait()
                fallidos_queue.put((pos, ruc, "No se pudo iniciar Chromium"))
                ruc_queue.task_done()
            except queue.Empty:
                break
        return

    # ── Loop principal del worker ─────────────────────────────────
    try:
        while not cancelar_flag():
            # Tomar el siguiente RUC de la cola (espera hasta 2 s)
            try:
                pos_global, ruc, cache_previo = ruc_queue.get(timeout=2)
            except queue.Empty:
                break   # Cola vacía → el worker terminó

            # ── Validación local (sin llamada a red) ──────────────
            error_local = _validar_ruc_local(ruc)
            if error_local:
                log("warn", f"RUC {ruc} — INVALIDO LOCAL: {error_local}")
                fila_inv = _armar_fila_invalida(ruc, error_local)
                with db_lock:
                    db_guardar(db_conn, ruc, {
                        FLAG_INVALIDO:      True,
                        ETAPA_RUC:          {"error_etapa": error_local},
                        ETAPA_DEUDA:        {"error_etapa": "RUC invalido"},
                        ETAPA_TRABAJADORES: {"error_etapa": "RUC invalido"},
                    })
                result_queue.put({"pos": pos_global, "fila": fila_inv})
                ruc_queue.task_done()
                continue

            # ── Bucle de reintentos ───────────────────────────────
            for intento in range(1, MAX_REINTENTOS + 1):
                if cancelar_flag():
                    break

                # Verificar que Chromium sigue vivo; reiniciar si es necesario
                if not _es_driver_vivo(driver):
                    driver_reinicios += 1
                    if driver_reinicios > MAX_DRIVER_REINICIOS:
                        log("error", "Chromium caido sin reinicios disponibles.")
                        fallidos_queue.put((pos_global, ruc, "Driver muerto"))
                        ruc_queue.task_done()
                        return
                    if not iniciar_o_reiniciar("Chromium no responde"):
                        fallidos_queue.put((pos_global, ruc, "No se pudo reiniciar Chromium"))
                        ruc_queue.task_done()
                        break

                try:
                    # En reintento N>1, aplicar backoff exponencial
                    if intento > 1:
                        espera = ESPERA_REINTENTO_BASE * (2 ** (intento - 2))
                        log("warn",
                            f"RUC {ruc} — reintento {intento}/{MAX_REINTENTOS} "
                            f"(pausa {espera:.0f}s)...")
                        time.sleep(espera)

                    # ── Consulta de las 3 etapas ──────────────────
                    fila       = consultar_ruc_completo(driver, ruc, cache_previo)
                    completado = fila.get("completado", False)
                    nombre     = fila.get("Nombre", "-")
                    raw        = fila.get("_raw", {})

                    # Persistir en SQLite (bajo lock para evitar conflictos)
                    with db_lock:
                        db_guardar(db_conn, ruc, raw)

                    # Log del resultado
                    if completado:
                        deuda = fila.get("Deuda_Tiene", "-")
                        log(
                            "error" if deuda == "SI" else "ok",
                            f"RUC {ruc} | {nombre} — "
                            + ("CON DEUDA COACTIVA" if deuda == "SI"
                               else f"OK (deuda:{deuda})")
                        )
                    else:
                        log("warn",
                            f"RUC {ruc} | {nombre} — PARCIAL: {fila.get('Etapas_Fallidas')}")
                        cache_previo = raw   # guardar lo obtenido para próximo intento

                    # Bajar el contador WAF al tener éxito
                    with waf_counter_lock:
                        waf_counter[0] = max(0, waf_counter[0] - 1)

                    result_queue.put({"pos": pos_global, "fila": fila})

                    # Salir del loop de reintentos si todo salió bien
                    if completado or intento == MAX_REINTENTOS:
                        break

                except RUCInvalidoException as e:
                    # SUNAT dice que el RUC no existe → no reintentar
                    log("warn", f"RUC {ruc} — INVALIDO SUNAT: {str(e)[:80]}")
                    fila_inv = _armar_fila_invalida(ruc, f"RUC invalido: {str(e)[:100]}")
                    with db_lock:
                        db_guardar(db_conn, ruc, {
                            FLAG_INVALIDO:      True,
                            ETAPA_RUC:          {"error_etapa": str(e)[:100]},
                            ETAPA_DEUDA:        {"error_etapa": "RUC invalido"},
                            ETAPA_TRABAJADORES: {"error_etapa": "RUC invalido"},
                        })
                    result_queue.put({"pos": pos_global, "fila": fila_inv})
                    break

                except WAFBlockException:
                    # SUNAT bloqueó este worker (o quizás toda la IP)
                    with waf_counter_lock:
                        waf_counter[0] += 1
                        activos = waf_counter[0]

                    if activos / max(total_workers, 1) >= UMBRAL_WAF_IP:
                        # Bloqueo de IP → pausa larga global
                        log("error",
                            f"IP POSIBLEMENTE BLOQUEADA ({activos}/{total_workers}). "
                            f"Pausa {PAUSA_IP_BLOQUEADA:.0f}s...")
                        time.sleep(PAUSA_IP_BLOQUEADA)
                        with waf_counter_lock:
                            waf_counter[0] = 0
                    else:
                        # Bloqueo individual → pausa corta
                        log("warn",
                            f"RUC {ruc} — WAF (intento {intento}) "
                            f"pausa {ESPERA_WAF:.0f}s...")
                        time.sleep(ESPERA_WAF)

                    if intento == MAX_REINTENTOS:
                        log("warn", f"RUC {ruc} — WAF agotado, marcando fallido.")
                        fallidos_queue.put((pos_global, ruc,
                                            f"Bloqueado WAF tras {MAX_REINTENTOS} intentos"))
                        with waf_counter_lock:
                            waf_counter[0] = max(0, waf_counter[0] - 1)
                        break

                except (PlaywrightError, PlaywrightTimeoutError) as e:
                    # Error de red o de Playwright (no WAF)
                    err_str = str(e)[:120]
                    log("warn", f"RUC {ruc} — Playwright (intento {intento}): {err_str}")
                    if not _es_driver_vivo(driver):
                        driver_reinicios += 1
                        if driver_reinicios <= MAX_DRIVER_REINICIOS:
                            iniciar_o_reiniciar("excepcion Playwright")
                    if intento == MAX_REINTENTOS:
                        fallidos_queue.put((pos_global, ruc, f"Error red: {err_str[:80]}"))

                except Exception as e:
                    # Cualquier otro error inesperado
                    err_str = str(e)[:120]
                    log("error",
                        f"RUC {ruc} — error inesperado (intento {intento}): {err_str}")
                    if intento == MAX_REINTENTOS:
                        fallidos_queue.put((pos_global, ruc, f"Error: {err_str[:80]}"))

            ruc_queue.task_done()
            time.sleep(_pausa_entre_rucs())    # pausa humana antes del siguiente RUC

    finally:
        # Siempre cerrar Chromium al salir, incluso si hubo excepción
        cerrar_driver_seguro(driver)
        log("dim", "Chromium cerrado.")


# ══════════════════════════════════════════════════════════════════
#  11  EXPORTACIÓN A EXCEL
#  ─────────────────────────────────────────────────────────────────
#  Convierte la lista de dicts de resultados en un archivo .xlsx
#  formateado con colores por estado (deuda/sin deuda/parcial/etc.)
#
#  Estructura del Excel:
#    Hoja "CONSULTA_SUNAT":  Una fila por RUC consultado.
#    Hoja "Consultas_Fallidas" (si aplica): RUCs que agotaron reintentos.
#
#  cargar_rucs_desde_excel() se incluye aquí por cohesión: es la
#  función que lee el Excel de entrada (hoja "ruc_unificado").
# ══════════════════════════════════════════════════════════════════

# ── Definición de columnas del Excel de salida ────────────────────
# Cada tupla es (clave_en_dict_resultado, encabezado_en_excel).
COLUMNAS_EXCEL = [
    ("RUC",                          "RUC"),
    ("Nombre",                       "Proveedor / Razon Social"),
    ("Tipo_Contribuyente",           "Tipo Contribuyente"),
    ("Nombre_Comercial",             "Nombre Comercial"),
    ("Fecha_Inscripcion",            "Fecha Inscripcion"),
    ("Fecha_Inicio_Actividades",     "Fecha Inicio Actividades"),
    ("Estado_Contribuyente",         "Estado Contribuyente"),
    ("Condicion_Contribuyente",      "Condicion Contribuyente"),
    ("Domicilio_Fiscal",             "Domicilio Fiscal"),
    ("Actividad_Economica_Principal","Actividad Economica Principal"),
    ("Actividades_Economicas",       "Actividades Economicas (todas)"),
    ("Deuda_Tiene",                  "Tiene Deuda Coactiva?"),
    ("Deuda_Monto",                  "Deuda Coactiva S/ (ultimo)"),
    ("Deuda_Periodo",                "Periodo Tributario (ultimo)"),
    ("Deuda_Fecha_Inicio_Cobranza",  "Fecha Inicio Cobranza (ultimo)"),
    ("Deuda_Entidad",                "Entidad Asociada Deuda (ultimo)"),
    ("Trab_Tiene",                   "Declara Trabajadores?"),
    ("Trab_Periodo",                 "Periodo Trabajadores (ultimo)"),
    ("Trab_Num_Trabajadores",        "Num Trabajadores (ultimo)"),
    ("Trab_Num_Pensionistas",        "Num Pensionistas (ultimo)"),
    ("Trab_Num_Prestadores",         "Num Prestadores Servicio (ultimo)"),
    ("Error",                        "Error / Observaciones"),
]

# Anchos de columna en Excel (letras A-W → ancho en caracteres)
ANCHOS_EXCEL = {
    "A": 16, "B": 40, "C": 22, "D": 22, "E": 18, "F": 22,
    "G": 20, "H": 20, "I": 44, "J": 36, "K": 50,
    "L": 18, "M": 22, "N": 18, "O": 26, "P": 34,
    "Q": 18, "R": 20, "S": 20, "T": 20, "U": 26, "V": 55,
}

# Colores de relleno por estado de la fila (solo si openpyxl disponible)
FILLS_EXCEL = {
    "CON_DEUDA": PatternFill("solid", fgColor="FFC7CE"),   # rojo claro
    "SIN_DEUDA": PatternFill("solid", fgColor="E2EFDA"),   # verde claro
    "INVALIDO":  PatternFill("solid", fgColor="D9D9D9"),   # gris
    "ERROR":     PatternFill("solid", fgColor="FCE4D6"),   # naranja claro
    "PARCIAL":   PatternFill("solid", fgColor="FFEB9C"),   # amarillo
    "NORMAL":    PatternFill("solid", fgColor="FFFFFF"),   # blanco
} if DEPS_OK else {}


def _escribir_encabezados(ws, encabezados, color_fondo):
    """
    Escribe la fila de encabezados de una hoja Excel con estilo.

    Aplica fondo de color, fuente blanca en negrita, alineación
    centrada y borde fino a todas las celdas del encabezado.

    Args:
        ws:           Worksheet de openpyxl.
        encabezados:  Lista de strings con los títulos de columna.
        color_fondo:  Color hex sin '#' para el fondo (ej. "1A3A5A").
    """
    fill  = PatternFill("solid", fgColor=color_fondo)
    font  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    for ci, nombre in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=ci, value=nombre)
        c.fill = fill; c.font = font; c.alignment = align; c.border = borde


def exportar_a_excel(todas_filas, fallidos, ruta_dest):
    """
    Genera el archivo Excel de resultados y lo guarda en ruta_dest.

    Crea:
    - Hoja "CONSULTA_SUNAT" con una fila por cada RUC.
      Filas coloreadas según estado:
        · Rojo    → tiene deuda coactiva
        · Verde   → sin deuda
        · Gris    → RUC inválido
        · Naranja → completado con errores
        · Amarillo → resultado parcial (alguna etapa falló)
        · Blanco  → estado desconocido
    - Hoja "Consultas_Fallidas" (solo si hay fallidos).

    Args:
        todas_filas: Lista de dicts planos de resultado (9).
        fallidos:    Lista de tuplas (ruc, motivo) de fallidos definitivos.
        ruta_dest:   Ruta del archivo .xlsx a crear.

    Returns:
        ruta_dest (str) para confirmación.
    """
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws      = wb.create_sheet("CONSULTA_SUNAT")
    keys    = [k for k, _ in COLUMNAS_EXCEL]
    headers = [h for _, h in COLUMNAS_EXCEL]

    _escribir_encabezados(ws, headers, "1A3A5A")
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"    # Fila 1 fija al hacer scroll

    # Establecer anchos de columna
    letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    for i in range(len(COLUMNAS_EXCEL)):
        col = letras[i] if i < 26 else letras[i // 26 - 1] + letras[i % 26]
        ws.column_dimensions[col].width = ANCHOS_EXCEL.get(col, 18)

    font_d  = Font(name="Arial", size=8)
    align_d = Alignment(horizontal="left", vertical="center", wrap_text=True)
    borde_d = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # ── Escribir filas de datos ───────────────────────────────────
    for fi, row in enumerate(todas_filas, 2):
        es_inv     = row.get("es_invalido", False)
        deuda      = row.get("Deuda_Tiene", "").upper()
        err        = row.get("Error", "")
        completado = row.get("completado", True)

        # Determinar color de la fila según su estado
        if es_inv:
            fill_row = FILLS_EXCEL["INVALIDO"]
        elif err and not completado:
            fill_row = FILLS_EXCEL["PARCIAL"]
        elif err:
            fill_row = FILLS_EXCEL["ERROR"]
        elif deuda == "SI":
            fill_row = FILLS_EXCEL["CON_DEUDA"]
        elif deuda == "NO":
            fill_row = FILLS_EXCEL["SIN_DEUDA"]
        else:
            fill_row = FILLS_EXCEL["NORMAL"]

        for ci, key in enumerate(keys, 1):
            val = row.get(key, "-") or "-"
            if isinstance(val, list):
                val = ", ".join(val) if val else "-"
            c = ws.cell(row=fi, column=ci, value=val)
            c.font = font_d; c.alignment = align_d
            c.border = borde_d; c.fill = fill_row

    # ── Hoja de fallidos (solo si los hay) ───────────────────────
    if fallidos:
        ws_f = wb.create_sheet("Consultas_Fallidas")
        _escribir_encabezados(ws_f, ["RUC", "Motivo"], "843C0C")
        fill_f  = PatternFill("solid", fgColor="FCE4D6")
        font_f  = Font(name="Arial", size=8)
        align_f = Alignment(horizontal="left", vertical="center", wrap_text=True)
        borde_f = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        for fi, (ruc, motivo) in enumerate(fallidos, 2):
            for ci, val in enumerate([ruc, motivo], 1):
                c = ws_f.cell(row=fi, column=ci, value=val)
                c.fill = fill_f; c.font = font_f
                c.alignment = align_f; c.border = borde_f
        ws_f.column_dimensions["A"].width = 16
        ws_f.column_dimensions["B"].width = 70
        ws_f.freeze_panes = "A2"

    wb.save(ruta_dest)
    wb.close()
    return ruta_dest


def cargar_rucs_desde_excel(ruta):
    """
    Lee la lista de RUCs desde la hoja "ruc_unificado" del Excel de entrada.

    Lee la columna A desde la fila 2 en adelante (la fila 1 es encabezado).
    Convierte cada celda a string, elimina decimales de Excel (ej. "20123.0")
    y filtra celdas vacías.

    Args:
        ruta: Ruta del archivo .xlsx o .xlsm.

    Returns:
        Lista de strings con los RUCs (sin validar formato; eso lo hace
        _validar_ruc_local() más adelante).

    Raises:
        ValueError: Si no existe la hoja "ruc_unificado" o no tiene datos.
    """
    keep_vba = ruta.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(ruta, read_only=True, keep_vba=keep_vba)
    try:
        if "ruc_unificado" not in wb.sheetnames:
            raise ValueError("No se encontro la hoja 'ruc_unificado'.")
        ws   = wb["ruc_unificado"]
        rucs = []
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            val = row[0]
            if val is None:
                continue
            # Eliminar ".0" que Excel añade a números (20123456789.0 → 20123456789)
            s = str(val).strip().split(".")[0].strip()
            if s:
                rucs.append(s)
        if not rucs:
            raise ValueError("La columna A de 'ruc_unificado' no contiene RUCs.")
        return rucs
    finally:
        wb.close()


# ══════════════════════════════════════════════════════════════════
#  12  INTERFAZ GRÁFICA — App (Tkinter)
#  ─────────────────────────────────────────────────────────────────
#  Clase principal de la aplicación.  Hereda de tk.Tk para ser
#  la ventana raíz.
#
#  Responsabilidades:
#    - Construir todos los widgets (12a _build_ui y sub-métodos)
#    - Gestionar el ciclo de selección → proceso → exportación
#    - Comunicarse con los workers vía queues (poll cada 150 ms)
#    - Mostrar progreso, ETA y resultados en tiempo real
#
#  Métodos agrupados por función:
#    _build_ui / _build_tabla / _build_log  → construcción de widgets
#    _seleccionar_archivo / _cargar_archivo  → flujo de apertura de archivo
#    _iniciar / _cancelar / _ejecutar        → flujo de consulta
#    _exportar                               → flujo de exportación
#    _poll                                   → loop de actualización de UI
#    _insertar_resultado_ui / _insertar_fila_tabla  → actualizar tabla
#    _aplicar_filtro / _fila_coincide        → búsqueda en tabla
#    _log / _log_direct / _update_estado    → mensajes al usuario
#    _avanzar_contador / _actualizar_estado_conteo / _set_progress → stats
#    on_close                                → limpieza al cerrar
# ══════════════════════════════════════════════════════════════════

class App(tk.Tk):
    """
    Ventana principal de la aplicación Consulta Masiva SUNAT.

    Gestiona toda la interacción del usuario: selección de archivo,
    inicio/cancelación de consultas, visualización de resultados
    en tiempo real y exportación a Excel.
    """

    # ── Paleta de colores (tema oscuro) ───────────────────────────
    BG      = "#0E0808"    # fondo general (casi negro rojizo)
    PANEL   = "#180D0D"    # paneles secundarios
    PANEL2  = "#201010"    # paneles de menor jerarquía
    ACCENT  = "#E05020"    # color de acento (naranja-rojo)
    ACCENT2 = "#7B1C1C"    # acento oscuro (barra de título)
    TEXT    = "#F0D8D8"    # texto principal
    DIM     = "#8A9BAD"    # texto secundario / inactivo
    OK      = "#229954"    # verde para éxito / botón Iniciar
    ERR     = "#C0392B"    # rojo para errores
    BORDER  = "#2A1212"    # bordes y separadores
    SEL     = "#2A1010"    # fondo de selección en Treeview

    # ── Columnas de la tabla de resultados ────────────────────────
    COL_KEYS    = [
        "RUC", "Nombre", "Estado_Contribuyente", "Condicion_Contribuyente",
        "Deuda_Tiene", "Deuda_Monto", "Trab_Tiene", "Trab_Num_Trabajadores", "Error",
    ]
    COL_HEADERS = [
        "RUC", "Proveedor", "Estado", "Condicion",
        "Deuda?", "Monto Deuda", "Trabajadores?", "Num Trab.", "Error / Etapas Fallidas",
    ]
    COL_WIDTHS = [110, 220, 80, 80, 65, 100, 90, 80, 220]

    # ─────────────────────────────────────────────────────────────
    # INICIALIZACIÓN
    # ─────────────────────────────────────────────────────────────

    def __init__(self):
        """
        Inicializa la ventana, variables de estado y construye la UI.

        Variables de estado principales:
            _ruta_real:      Ruta del archivo Excel seleccionado.
            en_proceso:      True mientras los workers están corriendo.
            _cancelar_flag:  True cuando el usuario pidió cancelar.
            _todas_filas:    Lista acumulada de dicts de resultado.
            _fallidos:       Lista de (ruc, motivo) de fallidos definitivos.
            _db_conn:        Conexión SQLite compartida con los workers.
            _db_lock:        Lock para serializar accesos a _db_conn.
        """
        super().__init__()
        self.title("Consulta Masiva SUNAT")
        self.geometry("1340x800")
        self.minsize(900, 540)
        self.configure(bg=self.BG)

        # ── Variables de estado del archivo ───────────────────────
        self.ruta_excel     = tk.StringVar(value="Ningun archivo seleccionado")
        self._ruta_real     = ""      # ruta completa del Excel seleccionado

        # ── Variables de estado del proceso ───────────────────────
        self.en_proceso     = False   # True mientras los workers corren
        self._cancelar_flag = False   # señal de cancelación para los workers
        self._log_visible   = False   # estado del panel de log (expandido/colapsado)

        # ── Acumuladores de resultados ────────────────────────────
        self._todas_filas     = []    # lista de dicts resultado (uno por RUC)
        self._fallidos        = []    # lista de (ruc, motivo) fallidos definitivos
        self._total_rucs      = 0     # total de RUCs en el archivo
        self._proc_rucs       = 0     # RUCs procesados en la sesión actual
        self._con_deuda       = 0     # contador de RUCs con deuda coactiva
        self._exportado       = False # True si se exportó al menos una vez
        self._ruta_export_dir = ""    # directorio del último export (para re-abrir)

        # ── Opciones de usuario ───────────────────────────────────
        self._auto_exportar = tk.BooleanVar(value=False)  # exportar al terminar
        self._usar_cache    = tk.BooleanVar(value=True)   # usar cache SQLite

        # ── ETA ───────────────────────────────────────────────────
        # Deque de tiempos (segundos) entre RUCs procesados
        self._tiempos_recientes = deque(maxlen=VENTANA_ETA)
        self._ultimo_ts         = None   # timestamp del último RUC procesado

        # ── Queues de comunicación con workers ────────────────────
        self._result_queue   = None   # resultados de RUCs completados
        self._log_queue      = None   # mensajes de log de los workers
        self._fallidos_queue = None   # RUCs fallidos definitivamente

        # ── Base de datos ─────────────────────────────────────────
        self._db_conn = None
        self._db_lock = threading.Lock()

        # ── Filtro de tabla ───────────────────────────────────────
        self._filtro_activo = ""

        self._build_ui()
        self._check_deps()
        self._poll()    # arranca el loop de actualización de UI

    # ─────────────────────────────────────────────────────────────
    # CONSTRUCCIÓN DE LA INTERFAZ
    # ─────────────────────────────────────────────────────────────

    def _build_ui(self):
        """
        Construye todos los widgets de la ventana principal.

        Jerarquía de widgets (de arriba a abajo):
            1. Barra de título (top)        — nombre de la app y contador
            2. Barra de controles (ctrl)    — botones y opciones
            3. Barra de estadísticas (stats) — leyenda de colores y rate
            4. Barra de progreso            — ttk.Progressbar
            5. Barra de estado + ETA        — texto de estado y tiempo estimado
            6. Barra de filtro              — campo de búsqueda en tabla
            7. Panel principal (PanedWindow):
               ├─ Tabla de resultados (Treeview)
               └─ Panel de log (colapsable)
        """
        # ── 1. Barra de título ────────────────────────────────────
        top = tk.Frame(self, bg=self.ACCENT2, height=44)
        top.pack(fill="x"); top.pack_propagate(False)
        tk.Label(
            top,
            text="  CONSULTA MASIVA SUNAT  v2.1  |  2 Workers",
            font=("Consolas", 11, "bold"), bg=self.ACCENT2, fg=self.TEXT,
        ).pack(side="left", padx=16, pady=10)
        # Contador "X/Y RUCs" en el extremo derecho de la barra
        self.lbl_contador = tk.Label(
            top, text="",
            font=("Consolas", 10), bg=self.ACCENT2, fg="#FFD060",
        )
        self.lbl_contador.pack(side="right", padx=16)

        # ── 2. Barra de controles ─────────────────────────────────
        ctrl = tk.Frame(self, bg=self.PANEL, height=50)
        ctrl.pack(fill="x"); ctrl.pack_propagate(False)

        def sep():
            """Insertar separador vertical entre grupos de controles."""
            tk.Frame(ctrl, bg=self.BORDER, width=1).pack(
                side="left", fill="y", pady=8, padx=2
            )

        # Botón "Abrir archivo"
        tk.Button(
            ctrl, text="  Abrir archivo",
            font=("Consolas", 9, "bold"), bg=self.PANEL2, fg=self.TEXT,
            activebackground=self.SEL, activeforeground=self.ACCENT,
            relief="flat", cursor="hand2", padx=12, pady=7,
            command=self._seleccionar_archivo,
        ).pack(side="left", padx=(10, 4), pady=8)

        # Etiqueta con nombre del archivo seleccionado
        tk.Label(
            ctrl, textvariable=self.ruta_excel,
            font=("Consolas", 8), bg=self.PANEL, fg=self.DIM,
            anchor="w", width=32,
        ).pack(side="left", padx=(0, 4))

        sep()

        # Indicador de workers (fijo en 2, no es un dropdown)
        tk.Label(
            ctrl, text="  2 Workers",
            font=("Consolas", 9, "bold"), bg=self.PANEL2, fg=self.DIM,
            padx=12, pady=7, relief="flat",
        ).pack(side="left", padx=(8, 4), pady=8)

        sep()

        # Botón INICIAR
        self.btn_iniciar = tk.Button(
            ctrl, text="  INICIAR",
            font=("Consolas", 10, "bold"), bg=self.OK, fg="#050D14",
            activebackground="#1A7040", relief="flat", cursor="hand2",
            padx=16, pady=7, command=self._iniciar,
        )
        self.btn_iniciar.pack(side="left", padx=(8, 3), pady=8)

        # Botón DETENER (cancelar el proceso en curso)
        self.btn_cancelar = tk.Button(
            ctrl, text="  DETENER",
            font=("Consolas", 10, "bold"), bg=self.BORDER, fg=self.DIM,
            relief="flat", cursor="hand2", padx=16, pady=7,
            state="disabled", command=self._cancelar,
        )
        self.btn_cancelar.pack(side="left", padx=(3, 8), pady=8)

        sep()

        # Botón "Exportar Excel"
        self.btn_exportar = tk.Button(
            ctrl, text="  Exportar Excel",
            font=("Consolas", 9, "bold"), bg=self.PANEL2, fg=self.DIM,
            activebackground=self.SEL, activeforeground=self.TEXT,
            relief="flat", cursor="hand2", padx=12, pady=7,
            state="disabled", command=self._exportar,
        )
        self.btn_exportar.pack(side="left", padx=(8, 4), pady=8)

        sep()

        # Checkbutton "Usar cache"
        tk.Checkbutton(
            ctrl, text="Usar BD/Cache",
            variable=self._usar_cache,
            font=("Consolas", 8), bg=self.PANEL, fg=self.DIM,
            selectcolor=self.PANEL2, activebackground=self.PANEL,
            activeforeground=self.TEXT, cursor="hand2",
            command=self._on_toggle_cache,
        ).pack(side="left", padx=(6, 2), pady=8)

        # Checkbutton "Auto-exportar al terminar"
        tk.Checkbutton(
            ctrl, text="Auto-exportar",
            variable=self._auto_exportar,
            font=("Consolas", 8), bg=self.PANEL, fg=self.DIM,
            selectcolor=self.PANEL2, activebackground=self.PANEL,
            activeforeground=self.TEXT, cursor="hand2",
        ).pack(side="left", padx=(6, 8), pady=8)

        # ── 3. Barra de estadísticas / leyenda ───────────────────
        stats = tk.Frame(self, bg=self.PANEL2, height=24)
        stats.pack(fill="x"); stats.pack_propagate(False)
        tk.Label(
            stats,
            text=(
                f"  Rojo=Con deuda  Verde=Sin deuda  Amarillo=Parcial  Gris=Invalido"
                f"  |  Pausa entre RUCs: 4-9s (+lectura 20%)"
                f"  |  Rate sensibles: 1/{RATE_LIMIT_SEG}s"
            ),
            font=("Consolas", 7, "bold"), bg=self.PANEL2, fg=self.DIM,
        ).pack(side="left", padx=10)
        # Contador de RUCs con deuda (se actualiza en tiempo real)
        self.lbl_deuda_count = tk.Label(
            stats, text="",
            font=("Consolas", 8, "bold"), bg=self.PANEL2, fg="#E74C3C",
        )
        self.lbl_deuda_count.pack(side="right", padx=14)

        # ── 4. Barra de progreso ──────────────────────────────────
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(
            "S.Horizontal.TProgressbar",
            troughcolor=self.PANEL, background=self.ACCENT,
            bordercolor=self.BG, lightcolor=self.ACCENT,
            darkcolor=self.ACCENT, thickness=4,
        )
        self.progress = ttk.Progressbar(
            self, style="S.Horizontal.TProgressbar", mode="determinate"
        )
        self.progress.pack(fill="x")

        # ── 5. Barra de estado + ETA ──────────────────────────────
        ef = tk.Frame(self, bg=self.BG)
        ef.pack(fill="x")
        self.lbl_estado = tk.Label(
            ef, text="  Selecciona un archivo para comenzar.",
            font=("Consolas", 8), bg=self.BG, fg=self.DIM,
            anchor="w", padx=10,
        )
        self.lbl_estado.pack(side="left", fill="x", expand=True, pady=(2, 0))
        # ETA ("Tiempo estimado restante") a la derecha
        self.lbl_eta = tk.Label(
            ef, text="",
            font=("Consolas", 8), bg=self.BG, fg=self.DIM,
            anchor="e", padx=14,
        )
        self.lbl_eta.pack(side="right", pady=(2, 0))

        # ── 6. Barra de filtro ────────────────────────────────────
        ff = tk.Frame(self, bg=self.PANEL2, height=30)
        ff.pack(fill="x"); ff.pack_propagate(False)
        tk.Label(
            ff, text="  Buscar:",
            font=("Consolas", 8), bg=self.PANEL2, fg=self.DIM,
        ).pack(side="left", padx=(8, 2), pady=4)
        self._filtro_var = tk.StringVar()
        self._filtro_var.trace_add("write", lambda *_: self._aplicar_filtro())
        tk.Entry(
            ff, textvariable=self._filtro_var,
            font=("Consolas", 9), bg=self.PANEL, fg=self.TEXT,
            insertbackground=self.TEXT, relief="flat", width=30,
        ).pack(side="left", padx=(0, 8), pady=4)
        tk.Label(
            ff, text="(RUC, nombre, estado o error)",
            font=("Consolas", 7), bg=self.PANEL2, fg=self.DIM,
        ).pack(side="left")
        tk.Button(
            ff, text="x limpiar",
            font=("Consolas", 7), bg=self.PANEL2, fg=self.DIM,
            relief="flat", cursor="hand2", padx=6,
            command=lambda: self._filtro_var.set(""),
        ).pack(side="left", padx=4)

        # ── 7. Panel principal (tabla + log) ─────────────────────
        self.main_pane = tk.PanedWindow(
            self, orient="vertical",
            bg=self.BORDER, sashwidth=5,
            sashrelief="flat", opaqueresize=True,
        )
        self.main_pane.pack(fill="both", expand=True, pady=(2, 0))

        tabla_frame = tk.Frame(self.main_pane, bg=self.BG)
        self._build_tabla(tabla_frame)
        self.main_pane.add(tabla_frame, stretch="always", minsize=200)

        self.log_outer = tk.Frame(self.main_pane, bg=self.PANEL)
        self._build_log(self.log_outer)
        self.main_pane.add(self.log_outer, stretch="never", minsize=26)

        # Inicialmente el log está colapsado (panel mínimo)
        self.after(120, lambda: self.main_pane.sash_place(0, 0, 9999))

    def _build_tabla(self, parent):
        """
        Construye el Treeview de resultados con scrollbars.

        El Treeview muestra COL_KEYS columnas con estilos de color
        por tag (con_deuda, sin_deuda, parcial, error_row, etc.).
        El doble clic en una fila copia el RUC al portapapeles.

        Args:
            parent: Frame contenedor donde se empaquetará la tabla.
        """
        # Encabezado de la tabla
        th = tk.Frame(parent, bg=self.PANEL2, height=26)
        th.pack(fill="x"); th.pack_propagate(False)
        tk.Label(
            th, text="  RESULTADOS — RUC + DEUDA COACTIVA + TRABAJADORES",
            font=("Consolas", 7, "bold"), bg=self.PANEL2, fg=self.DIM,
        ).pack(side="left", pady=5, padx=8)
        # Contador de filas visible en la esquina derecha del encabezado
        self.lbl_filas_tabla = tk.Label(
            th, text="0 filas",
            font=("Consolas", 8), bg=self.PANEL2, fg=self.DIM,
        )
        self.lbl_filas_tabla.pack(side="right", pady=5, padx=10)

        wrap = tk.Frame(parent, bg=self.BG)
        wrap.pack(fill="both", expand=True)

        # Estilo personalizado del Treeview
        style = ttk.Style()
        style.configure(
            "Dark.Treeview",
            background=self.PANEL, foreground=self.TEXT,
            fieldbackground=self.PANEL, rowheight=22,
            font=("Consolas", 8), borderwidth=0,
        )
        style.configure(
            "Dark.Treeview.Heading",
            background=self.ACCENT2, foreground="white",
            font=("Consolas", 8, "bold"), relief="flat",
        )
        style.map(
            "Dark.Treeview",
            background=[("selected", self.SEL)],
            foreground=[("selected", self.ACCENT)],
        )
        style.map("Dark.Treeview.Heading", background=[("active", "#9B2C2C")])
        style.layout("Dark.Treeview", [("Dark.Treeview.treearea", {"sticky": "nswe"})])

        self.tree = ttk.Treeview(
            wrap, columns=self.COL_KEYS,
            show="headings", style="Dark.Treeview",
            selectmode="extended",
        )
        for key, header, width in zip(self.COL_KEYS, self.COL_HEADERS, self.COL_WIDTHS):
            self.tree.heading(key, text=header, anchor="w")
            self.tree.column(
                key, width=width, minwidth=40, anchor="w",
                stretch=key in ("Nombre", "Error"),
            )

        # Tags de color por estado de la fila
        self.tree.tag_configure("con_deuda", background="#2A0808", foreground="#E74C3C")
        self.tree.tag_configure("sin_deuda", background="#071A0E", foreground="#2ECC71")
        self.tree.tag_configure("parcial",   background="#1A1800", foreground="#F1C40F")
        self.tree.tag_configure("error_row", background="#1A1000", foreground="#E67E22")
        self.tree.tag_configure("fallido",   background="#180505", foreground="#C0392B")
        self.tree.tag_configure("invalido",  background="#141414", foreground="#888888")
        self.tree.tag_configure("cacheado",  background="#0A0A1A", foreground="#5577BB")

        vsb = ttk.Scrollbar(wrap, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", self._copiar_ruc)

    def _build_log(self, parent):
        """
        Construye el panel de registro de actividad (log).

        Incluye:
        - Un encabezado con botón para expandir/colapsar el panel.
        - Un widget Text de solo lectura con colores por nivel de log.

        Niveles de log y colores:
            ok    → verde   (RUC OK)
            error → rojo    (RUC con deuda o error crítico)
            warn  → naranja (advertencia, reintento, WAF)
            info  → acento  (información general)
            dim   → gris    (detalles técnicos)

        Args:
            parent: Frame contenedor del panel de log.
        """
        # Encabezado colapsable
        lh = tk.Frame(parent, bg=self.BORDER, height=26)
        lh.pack(fill="x"); lh.pack_propagate(False)
        tk.Label(
            lh, text="  Registro de actividad",
            font=("Consolas", 8, "bold"), bg=self.BORDER, fg=self.DIM,
        ).pack(side="left", pady=5, padx=6)
        self.btn_toggle_log = tk.Button(
            lh, text="+ mostrar",
            font=("Consolas", 7, "bold"), bg=self.BORDER, fg=self.DIM,
            activebackground=self.PANEL2, activeforeground=self.TEXT,
            relief="flat", cursor="hand2", padx=8, pady=3,
            command=self._toggle_log,
        )
        self.btn_toggle_log.pack(side="right", padx=8, pady=3)

        # Cuerpo del log (oculto por defecto)
        self.log_body = tk.Frame(parent, bg=self.PANEL)
        self.txt_log = tk.Text(
            self.log_body, bg=self.PANEL, fg=self.TEXT,
            font=("Consolas", 8), relief="flat",
            state="disabled", wrap="word", height=8,
        )
        sb = tk.Scrollbar(self.log_body, command=self.txt_log.yview, bg=self.PANEL2)
        self.txt_log.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.txt_log.pack(fill="both", expand=True, padx=6, pady=4)

        # Colores por nivel de log
        for tag, color in [
            ("ok",    "#2ECC71"),
            ("error", "#E74C3C"),
            ("warn",  "#F39C12"),
            ("info",  self.ACCENT),
            ("dim",   self.DIM),
        ]:
            self.txt_log.tag_config(tag, foreground=color)

    # ─────────────────────────────────────────────────────────────
    # TOGGLE DE PANEL DE LOG
    # ─────────────────────────────────────────────────────────────

    def _toggle_log(self):
        """
        Alterna la visibilidad del panel de log.

        Cuando se expande, ajusta la posición del sash del PanedWindow
        para dar espacio al log sin ocultar la tabla.
        """
        self._log_visible = not self._log_visible
        if self._log_visible:
            self.log_body.pack(fill="both", expand=True)
            self.btn_toggle_log.config(text="- ocultar")
            self.main_pane.paneconfigure(self.log_outer, minsize=120)
            h = self.winfo_height()
            self.after(30, lambda: self.main_pane.sash_place(0, 0, max(200, h - 200)))
        else:
            self.log_body.pack_forget()
            self.btn_toggle_log.config(text="+ mostrar")
            self.main_pane.paneconfigure(self.log_outer, minsize=26)
            h = self.winfo_height()
            self.after(30, lambda: self.main_pane.sash_place(0, 0, h - 30))

    # ─────────────────────────────────────────────────────────────
    # FLUJO: SELECCIÓN Y CARGA DE ARCHIVO
    # ─────────────────────────────────────────────────────────────

    def _on_toggle_cache(self):
        """
        Callback cuando el usuario cambia el estado del checkbutton "Usar BD/Cache".

        Informa al usuario mediante el log del efecto del cambio.
        """
        if self._usar_cache.get():
            self._log("BD/Cache activada.", "dim")
        else:
            self._log("BD/Cache desactivada — se consultará todo desde cero.", "warn")

    def _seleccionar_archivo(self):
        """
        Abre un diálogo de archivo para seleccionar el Excel de entrada.

        No hace nada si ya hay un proceso en curso.
        Solo acepta .xlsx y .xlsm.
        """
        if self.en_proceso:
            return
        ruta = filedialog.askopenfilename(
            title="Seleccionar Excel con RUCs",
            filetypes=[("Archivos Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")],
        )
        if ruta:
            self._cargar_archivo(ruta)

    def _cargar_archivo(self, ruta):
        """
        Lee el archivo Excel y muestra resumen en la barra de estado.

        Lee la cantidad de RUCs del Excel y consulta el cache SQLite
        para mostrar cuántos ya están procesados (completos/parciales/inválidos).

        Args:
            ruta: Ruta del archivo Excel seleccionado.
        """
        self._ruta_real = ruta
        nombre = os.path.basename(ruta)
        if len(nombre) > 32:
            nombre = nombre[:29] + "..."
        self.ruta_excel.set(nombre)
        try:
            rucs = cargar_rucs_desde_excel(ruta)
            n    = len(rucs)
            # Intentar leer el estado del cache sin modificarlo
            try:
                conn = db_inicializar(ruta)
                proc = db_cargar_procesados(conn)
                conn.close()
                n_inv  = sum(1 for d in proc.values() if d.get(FLAG_INVALIDO))
                n_comp = sum(
                    1 for d in proc.values()
                    if not d.get(FLAG_INVALIDO)
                    and all(
                        not (d.get(e) or {}).get("error_etapa")
                        for e in [ETAPA_RUC, ETAPA_DEUDA, ETAPA_TRABAJADORES]
                    )
                )
                n_parc = len(proc) - n_inv - n_comp
            except Exception:
                n_inv = n_comp = n_parc = 0

            if not self._usar_cache.get():
                msg = f"  {n} RUCs — cache desactivado."
            elif n_comp or n_parc or n_inv:
                msg = (
                    f"  {n} RUCs | {n_comp} completos | "
                    f"{n_parc} parciales | {n_inv} invalidos | "
                    f"{n - n_comp - n_inv} pendientes"
                )
            else:
                msg = f"  {n} RUCs — listo para consultar"

            self._update_estado(msg)
            self._log(f"Archivo: {os.path.basename(ruta)} — {n} RUCs", "info")

        except Exception as e:
            self._update_estado(f"  Error: {e}")
            self._log(str(e), "error")
            self._ruta_real = ""

    # ─────────────────────────────────────────────────────────────
    # FLUJO: INICIO, CANCELACIÓN Y EJECUCIÓN
    # ─────────────────────────────────────────────────────────────

    def _iniciar(self):
        """
        Valida precondiciones y lanza el proceso de consulta masiva.

        Pasos:
        1. Verificar que las dependencias estén disponibles.
        2. Verificar que hay un archivo seleccionado.
        3. Limpiar la UI de la sesión anterior.
        4. Lanzar _ejecutar() en un hilo daemon.
        """
        if not DEPS_OK:
            messagebox.showerror("Dependencias faltantes", "\n".join(MISSING))
            return
        if not self._ruta_real or not os.path.exists(self._ruta_real):
            messagebox.showwarning("Sin archivo", "Selecciona un archivo Excel primero.")
            return
        if self.en_proceso:
            return

        # Limpiar estado de sesión anterior
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._todas_filas       = []
        self._fallidos          = []
        self._proc_rucs         = 0
        self._total_rucs        = 0
        self._con_deuda         = 0
        self._exportado         = False
        self._tiempos_recientes = deque(maxlen=VENTANA_ETA)
        self._ultimo_ts         = None
        self._filtro_var.set("")
        self.en_proceso         = True
        self._cancelar_flag     = False

        # Actualizar estado de botones
        self.btn_iniciar.config(state="disabled", bg="#1A2A1A", fg="#3A6A3A")
        self.btn_cancelar.config(state="normal", bg="#2A1010", fg=self.ERR,
                                  activebackground="#3A1515")
        self.btn_exportar.config(state="disabled", bg=self.PANEL2, fg=self.DIM)
        self.lbl_deuda_count.config(text="")
        self.lbl_eta.config(text="")

        threading.Thread(
            target=self._ejecutar, args=(self._ruta_real,), daemon=True
        ).start()

    def _cancelar(self):
        """
        Señala a los workers que deben detenerse.

        Activa _cancelar_flag; los workers verifican esta flag
        entre RUCs y en el bucle de reintentos.
        """
        self._cancelar_flag = True
        self._log("Cancelando...", "warn")
        self.btn_cancelar.config(state="disabled")

    def _ejecutar(self, ruta_origen):
        """
        Orquestador principal del proceso de consulta (corre en hilo separado).

        Pasos:
        1. Cargar RUCs del Excel.
        2. Inicializar SQLite (o limpiar si cache desactivado).
        3. Cargar RUCs ya procesados del cache y mostrarlos en la UI.
        4. Determinar RUCs pendientes (los que necesitan consulta).
        5. Crear Queue y lanzar NUM_WORKERS hilos worker.
        6. Esperar a que todos los workers terminen.
        7. Si auto-exportar activo, lanzar exportación.

        Args:
            ruta_origen: Ruta del archivo Excel de entrada.
        """
        try:
            rucs  = cargar_rucs_desde_excel(ruta_origen)
            total = len(rucs)

            # ── Inicializar cache SQLite ──────────────────────────
            usar_cache = self._usar_cache.get()
            try:
                self._db_conn = db_inicializar(ruta_origen)
                if usar_cache:
                    procesados_prev = db_cargar_procesados(self._db_conn)
                else:
                    db_limpiar(self._db_conn)
                    procesados_prev = {}
                    self._log("Cache limpiado.", "warn")
            except Exception as e:
                self._log(f"Warning cache SQLite: {e}", "warn")
                self._db_conn   = None
                procesados_prev = {}

            # ── Mostrar RUCs del cache en la tabla ────────────────
            n_cache = 0
            if usar_cache:
                for ruc, raw in procesados_prev.items():
                    fila = _armar_resultado_final(ruc, raw)
                    self._insertar_resultado_ui(fila, desde_cache=True)
                    n_cache += 1
                if n_cache:
                    self._log(f"{n_cache} RUCs cargados desde cache.", "dim")

            def _necesita_consulta(ruc):
                """
                Devuelve True si el RUC no está completo en el cache.
                Un RUC inválido NO necesita re-consulta.
                """
                if ruc not in procesados_prev:
                    return True
                raw = procesados_prev[ruc]
                if raw.get(FLAG_INVALIDO):
                    return False
                for etapa in [ETAPA_RUC, ETAPA_DEUDA, ETAPA_TRABAJADORES]:
                    d = raw.get(etapa)
                    if not d or d.get("error_etapa"):
                        return True
                return False

            # Filtrar solo los RUCs que necesitan consulta
            rucs_pendientes  = [(r, procesados_prev.get(r, {}))
                                for r in rucs if _necesita_consulta(r)]
            total_pendientes = len(rucs_pendientes)
            self._total_rucs = total

            self.after(0, lambda: self._set_progress(n_cache, total))
            self._update_estado(
                f"  Iniciando {NUM_WORKERS} Chromium(s)... "
                f"({total_pendientes} pendientes | "
                f"pausa 4-9s/RUC | rate {RATE_LIMIT_SEG}s)"
            )
            self._log(
                f"Total: {total} | Cache: {n_cache} | "
                f"Pendientes: {total_pendientes} | Workers: {NUM_WORKERS}",
                "info",
            )

            if not rucs_pendientes:
                self._log("Todos los RUCs estan completos en cache.", "ok")
                return

            # ── Crear queues e iniciar workers ─────────────────────
            self._result_queue   = queue.Queue()
            self._log_queue      = queue.Queue()
            self._fallidos_queue = queue.Queue()
            ruc_queue            = queue.Queue()

            for i, (ruc, cache) in enumerate(rucs_pendientes):
                ruc_queue.put((n_cache + i, ruc, cache))

            # Un user-agent distinto por worker (del pool shuffleado)
            waf_counter      = [0]
            waf_counter_lock = threading.Lock()
            ua_pool          = random.sample(USER_AGENTS, min(NUM_WORKERS, len(USER_AGENTS)))

            threads = []
            for wid in range(NUM_WORKERS):
                t = threading.Thread(
                    target=worker_procesar_rucs,
                    args=(
                        wid + 1, ruc_queue,
                        lambda: self._cancelar_flag,
                        self._log_queue, self._result_queue,
                        self._fallidos_queue,
                        waf_counter, waf_counter_lock, NUM_WORKERS,
                        self._db_conn, self._db_lock,
                        ua_pool[wid],
                    ),
                    daemon=True,
                )
                threads.append(t)
                t.start()

            # Esperar a que todos los workers terminen
            while any(t.is_alive() for t in threads):
                time.sleep(0.3)
                if self._cancelar_flag:
                    break
            for t in threads:
                t.join()

            self._log("Workers finalizados.", "dim")

            if self._auto_exportar.get() and not self._cancelar_flag:
                self._log("Auto-exportando...", "info")
                self.after(500, lambda: self._exportar(silencioso=True))

        except Exception as e:
            self._log(f"Error critico: {e}", "error")
            self._update_estado(f"  Error: {str(e)[:80]}")
        finally:
            self.en_proceso = False
            self.after(0, self._restaurar_botones_fin)

    def _restaurar_botones_fin(self):
        """
        Restaura el estado de los botones al finalizar el proceso.

        Llamado desde el hilo principal vía after() al terminar _ejecutar().
        Muestra un resumen en la barra de estado.
        """
        self.btn_iniciar.config(
            state="normal", bg=self.OK, fg="#050D14", activebackground="#1A7040"
        )
        self.btn_cancelar.config(state="disabled", bg=self.BORDER, fg=self.DIM)
        self.lbl_eta.config(text="")

        con_d  = self._con_deuda
        n_fall = len(self._fallidos)
        parc   = sum(
            1 for f in self._todas_filas
            if not f.get("completado", True) and not f.get("es_invalido", False)
        )
        estado = f"  Completado: {self._proc_rucs}/{self._total_rucs} RUCs"
        if con_d:  estado += f"  |  {con_d} con deuda"
        if parc:   estado += f"  |  {parc} parcial(es)"
        if n_fall: estado += f"  |  {n_fall} fallido(s)"
        self._update_estado(estado)

    # ─────────────────────────────────────────────────────────────
    # FLUJO: EXPORTACIÓN
    # ─────────────────────────────────────────────────────────────

    def _exportar(self, silencioso=False):
        """
        Guarda los resultados actuales en un archivo Excel.

        En modo normal: muestra diálogo de "Guardar como" y luego
        una alerta de resumen con estadísticas.

        En modo silencioso (auto-exportar): guarda directamente en
        el directorio del Excel de entrada con nombre automático.

        Args:
            silencioso: Si True, no muestra diálogos al usuario.

        Returns:
            True si se exportó correctamente, False si se canceló o falló.
        """
        if not self._todas_filas and not self._fallidos:
            if not silencioso:
                messagebox.showinfo("Sin datos", "No hay datos para exportar.")
            return False

        base_name  = (
            os.path.splitext(os.path.basename(self._ruta_real))[0]
            if self._ruta_real else "SUNAT"
        )
        timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_fn = f"{base_name}_SUNAT_{timestamp}.xlsx"
        init_dir   = (
            self._ruta_export_dir
            or (os.path.dirname(self._ruta_real) if self._ruta_real else "")
            or os.path.expanduser("~/Documents")
        )

        if silencioso:
            ruta_dest = os.path.join(init_dir, default_fn)
        else:
            ruta_dest = filedialog.asksaveasfilename(
                title="Guardar Excel de resultados",
                initialdir=init_dir, initialfile=default_fn,
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            )
            if not ruta_dest:
                return False

        self._ruta_export_dir = os.path.dirname(ruta_dest)
        try:
            exportar_a_excel(self._todas_filas, self._fallidos, ruta_dest)
            self._exportado = True
            self._log(f"Exportado: {ruta_dest}", "ok")
            if not silencioso:
                con_d = sum(1 for f in self._todas_filas if f.get("Deuda_Tiene") == "SI")
                sin_d = sum(1 for f in self._todas_filas if f.get("Deuda_Tiene") == "NO")
                parc  = sum(
                    1 for f in self._todas_filas
                    if not f.get("completado", True) and not f.get("es_invalido", False)
                )
                messagebox.showinfo(
                    "Exportacion completada",
                    f"Guardado en:\n{ruta_dest}\n\n"
                    f"Total: {len(self._todas_filas)} | "
                    f"Con deuda: {con_d} | Sin deuda: {sin_d}"
                    + (f"\nParciales: {parc}" if parc else "")
                    + (f"\nFallidos: {len(self._fallidos)}" if self._fallidos else ""),
                )
            return True
        except Exception as e:
            self._log(f"Error exportando: {e}", "error")
            if not silencioso:
                messagebox.showerror("Error", str(e))
            return False

    # ─────────────────────────────────────────────────────────────
    # LOOP DE ACTUALIZACIÓN DE UI (POLL)
    # ─────────────────────────────────────────────────────────────

    def _poll(self):
        """
        Loop de actualización de la UI (ejecutado en el hilo principal).

        Se llama a sí mismo cada 150 ms via after().
        Drena las tres queues de comunicación con los workers:
            - _log_queue      → mensajes de texto al widget log
            - _result_queue   → filas de resultado a la tabla
            - _fallidos_queue → RUCs fallidos definitivamente

        Procesa hasta N items por ciclo para no bloquear la UI.
        """
        try:
            # Mensajes de log (hasta 40 por ciclo)
            if self._log_queue:
                for _ in range(40):
                    try:
                        tag, msg = self._log_queue.get_nowait()
                        self._log_direct(msg, tag)
                    except queue.Empty:
                        break

            # Resultados de RUCs (hasta 20 por ciclo)
            if self._result_queue:
                for _ in range(20):
                    try:
                        data = self._result_queue.get_nowait()
                        fila = data.get("fila")
                        if fila:
                            self._insertar_resultado_ui(fila)
                    except queue.Empty:
                        break

            # RUCs fallidos definitivamente (hasta 10 por ciclo)
            if self._fallidos_queue:
                for _ in range(10):
                    try:
                        pos, ruc, motivo = self._fallidos_queue.get_nowait()
                        self._fallidos.append((ruc, motivo))
                        fila_f = _armar_fila_invalida(ruc, f"FALLIDO: {motivo}")
                        self._insertar_fila_tabla(fila_f, tag="fallido")
                        self._avanzar_contador()
                    except queue.Empty:
                        break
        except Exception:
            pass
        finally:
            # Reprogramar siempre, incluso si hubo excepción
            self.after(150, self._poll)

    # ─────────────────────────────────────────────────────────────
    # ACTUALIZACIÓN DE TABLA Y CONTADORES
    # ─────────────────────────────────────────────────────────────

    def _insertar_resultado_ui(self, fila: dict, desde_cache=False):
        """
        Determina el tag de color de una fila y la inserta en la tabla.

        También incrementa el contador de RUCs con deuda cuando aplica.

        Args:
            fila:        Dict plano de resultado (de _armar_resultado_final).
            desde_cache: Si True, usa el tag "cacheado" (azul oscuro)
                         y no llama a _avanzar_contador().
        """
        deuda      = (fila.get("Deuda_Tiene", "") or "").upper()
        err        = fila.get("Error", "")
        completado = fila.get("completado", True)
        es_inv     = fila.get("es_invalido", False)

        if es_inv:
            tag = "invalido"
        elif desde_cache:
            tag = "cacheado"
        elif not completado and err:
            tag = "parcial"
        elif err:
            tag = "error_row"
        elif deuda == "SI":
            tag = "con_deuda"
            self._con_deuda += 1
        elif deuda == "NO":
            tag = "sin_deuda"
        else:
            tag = "parcial"

        self._todas_filas.append(fila)
        self._insertar_fila_tabla(fila, tag=tag)
        if not desde_cache:
            self._avanzar_contador()

    def _avanzar_contador(self):
        """
        Incrementa el contador de RUCs procesados y actualiza la UI.

        Registra el tiempo transcurrido desde el último RUC para
        el cálculo del ETA, luego actualiza la barra de progreso
        y los labels de estado.
        """
        self._proc_rucs += 1
        ahora = time.time()
        if self._ultimo_ts is not None:
            self._tiempos_recientes.append(ahora - self._ultimo_ts)
        self._ultimo_ts = ahora
        self._set_progress(self._proc_rucs, self._total_rucs)
        self._actualizar_estado_conteo()

    def _actualizar_estado_conteo(self):
        """
        Actualiza los labels de estado, contador y ETA.

        El ETA se calcula promediando los últimos VENTANA_ETA tiempos
        y multiplicando por los RUCs restantes.
        """
        eta_txt = ""
        if self._tiempos_recientes and self._proc_rucs < self._total_rucs:
            velocidad = sum(self._tiempos_recientes) / len(self._tiempos_recientes)
            eta_seg   = velocidad * (self._total_rucs - self._proc_rucs)
            if eta_seg < 60:
                eta_txt = f"ETA ~{eta_seg:.0f}s"
            elif eta_seg < 3600:
                eta_txt = f"ETA ~{eta_seg / 60:.1f}min"
            else:
                eta_txt = f"ETA ~{eta_seg / 3600:.1f}h"
        self.lbl_eta.config(text=eta_txt, fg=self.DIM)
        self._update_estado(
            f"  Procesados: {self._proc_rucs}/{self._total_rucs}"
            f"  |  Con deuda: {self._con_deuda}"
        )
        self.lbl_contador.config(text=f"{self._proc_rucs}/{self._total_rucs} RUCs")
        if self._con_deuda:
            self.lbl_deuda_count.config(
                text=f"!  {self._con_deuda} RUC(s) con deuda coactiva"
            )

    def _insertar_fila_tabla(self, fila: dict, tag="sin_deuda"):
        """
        Inserta una fila en el Treeview y hace scroll hacia ella.

        También actualiza el contador de filas y habilita el botón
        de exportación si aún no estaba activo.

        Args:
            fila: Dict plano de resultado.
            tag:  Tag de color a aplicar a la fila.
        """
        vals = tuple(fila.get(k, "-") or "-" for k in self.COL_KEYS)
        self.tree.insert("", "end", values=vals, tags=(tag,))
        ch = self.tree.get_children()
        if ch:
            self.tree.see(ch[-1])    # scroll automático a la última fila
        n = len(ch)
        self.lbl_filas_tabla.config(text=f"{n} fila{'s' if n != 1 else ''}")
        self.btn_exportar.config(
            state="normal", bg=self.ACCENT, fg="#050D14",
            activebackground="#B04010",
        )

    # ─────────────────────────────────────────────────────────────
    # FILTRO DE TABLA
    # ─────────────────────────────────────────────────────────────

    def _aplicar_filtro(self):
        """
        Filtra las filas del Treeview según el texto del campo de búsqueda.

        Reconstruye el Treeview completo con solo las filas que coincidan.
        El filtro se aplica sobre: RUC, Nombre, Estado, Condicion,
        Deuda_Tiene, Deuda_Entidad y Error.

        Solo se reconstruye si el texto cambió respecto al filtro anterior.
        """
        texto = self._filtro_var.get().strip().lower()
        if texto == self._filtro_activo:
            return
        self._filtro_activo = texto

        # Limpiar tabla y repoblar con filas que coincidan
        for item in self.tree.get_children():
            self.tree.delete(item)

        for fila in self._todas_filas:
            if texto == "" or self._fila_coincide(fila, texto):
                deuda      = (fila.get("Deuda_Tiene", "") or "").upper()
                completado = fila.get("completado", True)
                es_inv     = fila.get("es_invalido", False)
                err        = fila.get("Error", "")
                if   es_inv:         tag = "invalido"
                elif not completado: tag = "parcial"
                elif err:            tag = "error_row"
                elif deuda == "SI":  tag = "con_deuda"
                elif deuda == "NO":  tag = "sin_deuda"
                else:                tag = "parcial"
                vals = tuple(fila.get(k, "-") or "-" for k in self.COL_KEYS)
                self.tree.insert("", "end", values=vals, tags=(tag,))

        n = len(self.tree.get_children())
        self.lbl_filas_tabla.config(
            text=f"{n} fila{'s' if n != 1 else ''}" + (" (filtro)" if texto else "")
        )

    def _fila_coincide(self, fila: dict, texto: str) -> bool:
        """
        Comprueba si una fila contiene el texto buscado en algún campo clave.

        Args:
            fila:  Dict plano de resultado.
            texto: Texto a buscar (ya en minúsculas).

        Returns:
            True si alguno de los campos clave contiene el texto.
        """
        campos = [
            fila.get(k, "") for k in [
                "RUC", "Nombre", "Estado_Contribuyente",
                "Condicion_Contribuyente", "Deuda_Tiene",
                "Deuda_Entidad", "Error",
            ]
        ]
        return any(texto in str(c).lower() for c in campos)

    # ─────────────────────────────────────────────────────────────
    # UTILIDADES DE UI
    # ─────────────────────────────────────────────────────────────

    def _copiar_ruc(self, event):
        """
        Callback de doble clic en la tabla: copia el RUC al portapapeles.

        Muestra un mensaje temporal en la barra de estado confirmando
        la copia, luego restaura el mensaje de estado anterior.

        Args:
            event: Evento de Tkinter (contiene coordenadas del clic).
        """
        item = self.tree.identify_row(event.y)
        if not item:
            return
        ruc = self.tree.set(item, "RUC")
        if not ruc or ruc == "-":
            return
        self.clipboard_clear()
        self.clipboard_append(ruc)
        self._update_estado(f"  Copiado: {ruc}")
        self.after(2000, lambda: self._update_estado(
            f"  Procesados: {self._proc_rucs}/{self._total_rucs}"
            if self._total_rucs
            else "  Selecciona un archivo para comenzar."
        ))

    def _log(self, texto, tag=""):
        """
        Encola un mensaje de log para ser mostrado en el hilo principal.

        Seguro para llamar desde cualquier hilo (usa after()).

        Args:
            texto: Mensaje a mostrar.
            tag:   Nivel de log: "ok", "error", "warn", "info", "dim".
        """
        self.after(0, lambda: self._log_direct(texto, tag))

    def _log_direct(self, texto, tag=""):
        """
        Inserta directamente un mensaje en el widget de log (hilo principal).

        Habilita temporalmente el Text (que está en estado "disabled"),
        inserta el texto con su tag de color y hace scroll al final.

        Args:
            texto: Mensaje a insertar.
            tag:   Tag de color configurado en _build_log().
        """
        try:
            self.txt_log.config(state="normal")
            self.txt_log.insert("end", texto + "\n", tag)
            self.txt_log.see("end")
            self.txt_log.config(state="disabled")
        except Exception:
            pass

    def _update_estado(self, texto):
        """
        Actualiza el texto de la barra de estado (hilo-seguro).

        Args:
            texto: Texto a mostrar en lbl_estado.
        """
        self.after(0, lambda: self.lbl_estado.config(text=texto, fg=self.TEXT))

    def _set_progress(self, actual, total):
        """
        Actualiza la barra de progreso (hilo-seguro).

        Args:
            actual: Valor actual (RUCs procesados).
            total:  Valor máximo (total de RUCs).
        """
        self.after(0, lambda: self.progress.configure(
            maximum=max(total, 1), value=actual
        ))

    # ─────────────────────────────────────────────────────────────
    # CIERRE DE LA APLICACIÓN
    # ─────────────────────────────────────────────────────────────

    def on_close(self):
        """
        Maneja el cierre de la ventana (WM_DELETE_WINDOW).

        Si hay un proceso en curso, pregunta al usuario si desea
        interrumpirlo.  Si hay datos sin exportar, ofrece exportarlos
        antes de cerrar.  Cierra la conexión SQLite antes de destruir.
        """
        # Si el proceso está corriendo, confirmar interrupción
        if self.en_proceso:
            if not messagebox.askyesno(
                "Consulta en curso",
                "Hay una consulta en curso.\n"
                "Cerrar de todas formas? Se perderan los resultados no exportados.",
            ):
                return
            self._cancelar_flag = True

        # Si hay datos sin exportar, ofrecer guardarlos
        hay_datos = bool(self._todas_filas or self._fallidos)
        if hay_datos and not self._exportado:
            resp = messagebox.askyesnocancel(
                "Exportar antes de cerrar?",
                f"Tienes {len(self._todas_filas)} resultado(s) sin exportar.\n\n"
                "Si -> Exportar y cerrar\n"
                "No -> Cerrar sin exportar\n"
                "Cancelar -> Volver",
            )
            if resp is None:
                return    # usuario canceló → no cerrar
            if resp is True:
                if not self._exportar():
                    return    # exportación cancelada → no cerrar

        # Cerrar SQLite si está abierto
        if self._db_conn:
            try:
                self._db_conn.close()
            except Exception:
                pass
        self.destroy()

    # ─────────────────────────────────────────────────────────────
    # VERIFICACIÓN DE DEPENDENCIAS AL INICIO
    # ─────────────────────────────────────────────────────────────

    def _check_deps(self):
        """
        Verifica las dependencias al arrancar y muestra el resultado en el log.

        Si faltan dependencias, las lista en el log con color de error.
        Si todo está OK, muestra la configuración actual del sistema.
        """
        if not DEPS_OK:
            self._log("Dependencias faltantes:", "warn")
            for m in MISSING:
                self._log(f"   {m}", "error")
        else:
            self._log("Dependencias OK — SUNAT", "ok")
            self._log(
                f"Workers fijos: {NUM_WORKERS}  |  "
                f"Timeout: {TIMEOUT_CARGA}s  |  "
                f"Reintentos: {MAX_REINTENTOS}",
                "info",
            )
            self._log(
                f"Pausa entre RUCs: 4-9s base + "
                f"lectura aleatoria (20% de RUCs: +8-18s)",
                "info",
            )
            self._log(
                f"Pausa entre etapas: {PAUSA_ETAPA_MIN}-{PAUSA_ETAPA_MAX}s  |  "
                f"Rate sensibles: 1 cada {RATE_LIMIT_SEG}s global",
                "info",
            )
            self._log(
                f"WAF individual: pausa {ESPERA_WAF}s  |  "
                f"IP bloqueada: {PAUSA_IP_BLOQUEADA}s",
                "dim",
            )
            self._log("Selecciona tu Excel con la hoja 'ruc_unificado'.", "dim")


# ══════════════════════════════════════════════════════════════════
#  13  PUNTO DE ENTRADA
#  ─────────────────────────────────────────────────────────────────
#  Solo se ejecuta cuando el archivo se lanza directamente
#  (no cuando se importa como módulo).
# ══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = App()
    # Registrar el método de cierre personalizado para manejar
    # datos sin exportar y proceso en curso antes de destruir la ventana.
    app.protocol("WM_DELETE_WINDOW", app.on_close)
    app.mainloop()